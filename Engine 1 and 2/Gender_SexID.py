import re
import sys
import time
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

from ethnic_taggerv3 import (
    normalize_text,
    get_column_texts,
    matches_any,
    is_negated,
    is_example_mention,
)
from constants import ASPIRATIONAL_PHRASES

from gender_constants import (
    # Gender identity labels
    GENDER_WOMEN_GIRLS, GENDER_MEN_BOYS, GENDER_TWO_SPIRIT,
    GENDER_OTHER, GENDER_MULTIPLE, GENDER_GENERAL_POP,
    # Gender output columns
    OUTPUT_GENDER, OUTPUT_GENDER_FLAG,
    # Gender term data
    IDENTITY_KEY_TO_LABEL, IDENTITY_KEY_SHORT_LABEL,
    GENDER_TERM_PATTERNS,
    BARE_QUEER_PATTERN, BARE_TRANS_PATTERN, GENDER_QUEER_CONTEXT,
    # Gender flags
    FLAG_SEX_TERM, FLAG_QUEER_AMBIGUOUS, FLAG_BARE_TRANS,
    FLAG_ASPIRATIONAL, FLAG_TWO_SPIRIT_INDIG, FLAG_NEGATION,
    # Sexual identity labels
    SEXUAL_2SLGBTQIA, SEXUAL_GENERAL_POP,
    # Sexual output columns
    OUTPUT_SEXUAL, OUTPUT_SEXUAL_FLAG,
    # Sexual term data
    SEXUAL_GENDER_DIVERSE_PATTERNS, SEXUAL_ORIENTATION_PATTERNS,
    # Sexual flags
    SFLAG_NEGATION, SFLAG_GENDER_TERM, SFLAG_ASPIRATIONAL,
)

"""
Gender_SexID.py
---------------
Gender Identity + Sexual Identity classifiers — standalone script.

Reads:   "Data Sheets/FR testing.xlsx"  →  sheet "Discretionary Funding Requests"
Writes:  four columns back into the same workbook (openpyxl header-lookup):
           Gender Id - FR9       Gender Classification Flag
           Sexual Id - FR10      Sexual Classification Flag

Architecture:
    All data (labels, patterns, flag strings) lives in gender_constants.py.
    Generic text helpers (normalize_text, is_negated, etc.) are imported
    from ethnic_taggerv3.py — not copied.

Public API (imported by generate_review_report.py):
    classify_gender(row)  → (label: str, flag: str)
    classify_sexual(row)  → (label: str, flag: str)
    GENDER_GENERAL_POP, GENDER_MULTIPLE
    SEXUAL_2SLGBTQIA, SEXUAL_GENERAL_POP

Pipeline — Gender:
    1. get_column_texts() → normalize → extract_gender_candidates()
    2. resolve_gender() : 0 keys → General Pop, 1 → that group, 2+ → Multiple

Pipeline — Sexual:
    1. same normalized text → extract_sexual_candidates()
    2. resolve_sexual() : any guarded signal → 2SLGBTQIA+, else General Pop

To Run:
    python Gender_SexID.py
"""

DATA_SHEET = "Discretionary Funding Requests"


# ===========================================================================
# GENDER IDENTITY — extraction + resolver
# ===========================================================================

def extract_gender_candidates(text):
    """
    Scan normalized text for gender-identity terms.

    Returns
    -------
    keys        : set[str]  — accepted identity keys
    flags_out   : set[str]  — annotation flag strings triggered by this text
    any_negated : bool      — True if any negation encountered (→ FLAG_NEGATION)
    """
    keys        = set()
    flags_out   = set()
    any_negated = False

    # Standard (unambiguous) patterns
    for pattern, identity_key, extra_flag_key in GENDER_TERM_PATTERNS:
        m = re.search(pattern, text, re.IGNORECASE)
        if not m:
            continue
        term = m.group(0)
        if is_negated(term, text):
            any_negated = True
            continue
        if is_example_mention(term, text):
            continue
        keys.add(identity_key)
        if extra_flag_key == "sex_term":
            flags_out.add(FLAG_SEX_TERM)

    # Bare "queer" — ambiguity check
    # \bqueer\b fires on "queer youth" AND on "gender queer community".
    # Suppress the ambiguity flag when "queer" is part of "gender queer".
    for m in re.finditer(BARE_QUEER_PATTERN, text, re.IGNORECASE):
        term = m.group(0)
        if is_negated(term, text) or is_example_mention(term, text):
            any_negated = True
            continue
        prefix = text[max(0, m.start() - 8) : m.start()].rstrip()
        if prefix.lower().endswith("gender") or prefix.lower().endswith("gender-"):
            continue  # already captured by the "gender queer" standard pattern
        keys.add("genderqueer")
        flags_out.add(FLAG_QUEER_AMBIGUOUS)

    # Bare "trans" — false-positive risk
    # \btrans\b cannot match inside "transgender" (no \b between 's' and 'g').
    for m in re.finditer(BARE_TRANS_PATTERN, text, re.IGNORECASE):
        term = m.group(0)
        if is_negated(term, text) or is_example_mention(term, text):
            any_negated = True
            continue
        keys.add("transgender")
        flags_out.add(FLAG_BARE_TRANS)

    return keys, flags_out, any_negated


def resolve_gender(keys, flags_set, any_negated, aspirational):
    """
    0 keys  → General Population (No specific gender served)
    1 key   → that key's output group label
    2+ keys → Multiple gender identities  (flag lists which identities)
    Annotation flags are appended; they never change the branch.
    """
    flag_parts = sorted(flags_set)

    if any_negated:
        flag_parts.append(FLAG_NEGATION)
    if aspirational:
        flag_parts.append(FLAG_ASPIRATIONAL)
    if "two_spirit" in keys:
        flag_parts.append(FLAG_TWO_SPIRIT_INDIG)

    if not keys:
        return GENDER_GENERAL_POP, "; ".join(flag_parts)

    if len(keys) == 1:
        return IDENTITY_KEY_TO_LABEL[next(iter(keys))], "; ".join(flag_parts)

    short_names = sorted(IDENTITY_KEY_SHORT_LABEL[k] for k in keys)
    flag_parts.insert(0, f"Multiple: {', '.join(short_names)}")
    return GENDER_MULTIPLE, "; ".join(flag_parts)


def classify_gender(row):
    """Public entry point — gender identity classification for a single row."""
    col_texts = get_column_texts(row)
    combined  = " ".join(t for t in col_texts if t.strip())
    if not combined.strip():
        return GENDER_GENERAL_POP, ""
    text = normalize_text(combined)
    keys, flags_set, any_negated = extract_gender_candidates(text)
    aspirational                  = matches_any(ASPIRATIONAL_PHRASES, text)
    return resolve_gender(keys, flags_set, any_negated, aspirational)


# ===========================================================================
# SEXUAL IDENTITY — extraction + resolver
# ===========================================================================

def extract_sexual_candidates(text):
    """
    Presence-only scan — any single guarded signal is enough.

    Returns
    -------
    found           : bool — at least one signal survived guards
    gender_term_only: bool — signals came exclusively from gender-diverse
                             patterns (not orientation terms / acronyms)
    any_negated     : bool — at least one negation encountered
    """
    found_gender_diverse = False
    found_orientation    = False
    any_negated          = False

    for pattern in SEXUAL_GENDER_DIVERSE_PATTERNS:
        m = re.search(pattern, text, re.IGNORECASE)
        if not m:
            continue
        term = m.group(0)
        if is_negated(term, text):
            any_negated = True
            continue
        if is_example_mention(term, text):
            continue
        found_gender_diverse = True

    for pattern in SEXUAL_ORIENTATION_PATTERNS:
        m = re.search(pattern, text, re.IGNORECASE)
        if not m:
            continue
        term = m.group(0)
        if is_negated(term, text):
            any_negated = True
            continue
        if is_example_mention(term, text):
            continue
        found_orientation = True

    found = found_gender_diverse or found_orientation
    gender_term_only = found_gender_diverse and not found_orientation
    return found, gender_term_only, any_negated


def resolve_sexual(found, gender_term_only, any_negated, aspirational):
    """
    found → 2SLGBTQIA+; else → General Population.
    Flags are annotations only — they never change the branch.
    """
    flag_parts = []

    if found:
        if gender_term_only:
            flag_parts.append(SFLAG_GENDER_TERM)
        if any_negated:
            flag_parts.append(SFLAG_NEGATION)
        if aspirational:
            flag_parts.append(SFLAG_ASPIRATIONAL)
        return SEXUAL_2SLGBTQIA, "; ".join(flag_parts)

    if any_negated:
        flag_parts.append(SFLAG_NEGATION)
    if aspirational:
        flag_parts.append(SFLAG_ASPIRATIONAL)
    return SEXUAL_GENERAL_POP, "; ".join(flag_parts)


def classify_sexual(row):
    """Public entry point — sexual identity classification for a single row."""
    col_texts = get_column_texts(row)
    combined  = " ".join(t for t in col_texts if t.strip())
    if not combined.strip():
        return SEXUAL_GENERAL_POP, ""
    text = normalize_text(combined)
    found, gender_term_only, any_negated = extract_sexual_candidates(text)
    aspirational                          = matches_any(ASPIRATIONAL_PHRASES, text)
    return resolve_sexual(found, gender_term_only, any_negated, aspirational)


# ===========================================================================
# main() — writes both classifiers to FR testing.xlsx in one workbook pass
# ===========================================================================

def main():
    start_time = time.time()

    SCRIPT_DIR       = Path(__file__).resolve().parent
    funding_filepath = SCRIPT_DIR.parent / "Data Sheets" / "FR testing.xlsx"

    print(f"Loading funding requests from: {funding_filepath}")
    try:
        data_df = pd.read_excel(funding_filepath, sheet_name=DATA_SHEET, dtype=str)
    except Exception as e:
        print(f"Error loading data sheet '{DATA_SHEET}' from '{funding_filepath}': {e}")
        sys.exit(1)

    print(f"Data rows: {len(data_df)}")

    for col in [OUTPUT_GENDER, OUTPUT_GENDER_FLAG, OUTPUT_SEXUAL, OUTPUT_SEXUAL_FLAG]:
        if col not in data_df.columns:
            data_df[col] = ""

    stats = {
        "gender_women_girls": 0, "gender_men_boys": 0, "gender_two_spirit": 0,
        "gender_other": 0, "gender_multiple": 0, "gender_general": 0,
        "sexual_2slgbtqia": 0, "sexual_general": 0,
        "flagged_gender": 0, "flagged_sexual": 0,
    }

    for idx, row in data_df.iterrows():
        g_label, g_flag = classify_gender(row)
        s_label, s_flag = classify_sexual(row)

        data_df.at[idx, OUTPUT_GENDER]      = g_label
        data_df.at[idx, OUTPUT_GENDER_FLAG] = g_flag
        data_df.at[idx, OUTPUT_SEXUAL]      = s_label
        data_df.at[idx, OUTPUT_SEXUAL_FLAG] = s_flag

        if   g_label == GENDER_WOMEN_GIRLS: stats["gender_women_girls"] += 1
        elif g_label == GENDER_MEN_BOYS:    stats["gender_men_boys"]    += 1
        elif g_label == GENDER_TWO_SPIRIT:  stats["gender_two_spirit"]  += 1
        elif g_label == GENDER_OTHER:       stats["gender_other"]       += 1
        elif g_label == GENDER_MULTIPLE:    stats["gender_multiple"]    += 1
        else:                               stats["gender_general"]     += 1

        if s_label == SEXUAL_2SLGBTQIA:     stats["sexual_2slgbtqia"]  += 1
        else:                               stats["sexual_general"]     += 1

        if g_flag: stats["flagged_gender"]  += 1
        if s_flag: stats["flagged_sexual"]  += 1

    # Write-back via openpyxl header-lookup (never assumes column positions)
    wb = load_workbook(funding_filepath)
    ws = wb[DATA_SHEET]
    headers = {cell.value: cell.column for cell in ws[1]}

    for col_name in [OUTPUT_GENDER, OUTPUT_GENDER_FLAG, OUTPUT_SEXUAL, OUTPUT_SEXUAL_FLAG]:
        if col_name not in headers:
            new_col = ws.max_column + 1
            ws.cell(row=1, column=new_col, value=col_name)
            headers[col_name] = new_col

    for i, (idx, _row) in enumerate(data_df.iterrows(), start=2):
        ws.cell(row=i, column=headers[OUTPUT_GENDER],      value=data_df.at[idx, OUTPUT_GENDER])
        ws.cell(row=i, column=headers[OUTPUT_GENDER_FLAG], value=data_df.at[idx, OUTPUT_GENDER_FLAG])
        ws.cell(row=i, column=headers[OUTPUT_SEXUAL],      value=data_df.at[idx, OUTPUT_SEXUAL])
        ws.cell(row=i, column=headers[OUTPUT_SEXUAL_FLAG], value=data_df.at[idx, OUTPUT_SEXUAL_FLAG])

    wb.save(funding_filepath)

    print("\nResults:")
    for k, v in stats.items():
        print(f"  {k}: {v}")
    print(f"\nOutput written to: {funding_filepath}")
    elapsed = time.time() - start_time
    print(f"Gender + Sexual classification completed in {elapsed:.1f} seconds.")


if __name__ == "__main__":
    main()
