import sys
import re
import pandas as pd
import time

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path

"""
ethnic_tagger_v3.py
--------------------
Rule-based ethnic origin classifier for ECF Discretionary
Funding Requests.

Reads:
    "Taxonomy - Definitions.xlsx" -> sheet "Ethnic and Cultural Origins"
    "Discretionary FR working -2025 (Oreva).xlsx" -> sheet "Discretionary Funding Requests"

Aggregates matches across all 4 input columns (Final_Project_Description,
Final_Summary_Description, Purpose, Funding Request Name), then resolves
using the case hierarchy below.

To Run:
    python ethnic_taggerv3.py "C:\\Users\\oadode\\OneDrive - Edmonton Community Foundation\\Desktop\\Discretionary FR Scripting\\ECF-Discretionary-FR-Scripting\\Taxonomy - Definitions.xlsx" "C:\\Users\\oadode\\OneDrive - Edmonton Community Foundation\\Desktop\\Discretionary FR Scripting\\ECF-Discretionary-FR-Scripting\\FR testing.xlsx"

Case coverage (see README):
    1.  Exact match (deepest taxonomy term)
    2.  Level 2 match (subregion)
    3.  Level 1 match (broad category)
    4.  Structured phrase ("South African" etc.) -> pattern rules
    5.  Country/nationality IN taxonomy -> handled by case 1-3
    6.  Country/nationality NOT in taxonomy -> country map
    7.  "from <country>" phrasing -> country map (phrase variant)
    8.  Multiple groups -> Multiple Ethnic and Cultural Origins
    9.  BIPOC (context-aware) -> Multiple Ethnic and Cultural Origins
    9b. Broad identity labels (Black, Arab, Jewish, etc.) -> Other Ethnic
        and Cultural Origins
    10. Known organization name lookup (Bent Arrow, Treaty 6, etc.)
        -> only consulted if result would otherwise be General Population
    11. "Grassroots" -> ethnic-indicating ONLY if paired with another
        ethnic keyword; otherwise ignored entirely
    12. General Population (fallback / no signal)

Also handles:
    - Context override ("beyond its original focus on X")
        - Historical reference ("historically served X")
        - Expansion Phrases ("")
    - Negation ("not targeting X", "does not serve X")
    - Aspirational language ("plans to expand to X") -> flagged for review
    - Example mentions ("such as X", "including X") -> not classified
    - Strict word-boundary matching (no "blacksmith" -> "black")
"""

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))  # Engine 1 and 2
import bootstrap 

# Semantic-similarity fallback (see semantic_fallback.py).
try:
    import semantic_fallback
    SEMANTIC_AVAILABLE = True
except ImportError:
    SEMANTIC_AVAILABLE = False

from constants import (
    MULTIPLE_ETHNIC, OTHER_ETHNIC, GENERAL_POP,
    BIPOC_KEYWORDS, AMBIGUOUS_EQUITY_WORDS, BROAD_IDENTITY_KEYWORDS,
    ALWAYS_MULTIPLE_COMPOUNDS, ORG_NAME_ETHNICITY_MAP,
    PATTERN_RULES, COUNTRY_REGION_MAP,
    EXPANSION_PHRASES, HISTORICAL_PHRASES, NEGATION_PHRASES,
    ETHNIC_ANNOTATION_NEGATION_PHRASES,
    ASPIRATIONAL_PHRASES, EXAMPLE_PHRASES, EXPERT_ROLE_PHRASES,
    SERVING_CONTEXT_WORDS, IDENTITY_PHRASE_REWRITES,
    DIRECTIONAL_AFRICAN_PREFIXES, CASE_13_PATTERNS,
    DEMONYM_SUFFIXES, NON_ETHNIC_LEADING_WORDS,
)

# =============
# CONFIGURATION 
# =============

TAXONOMY_SHEET = "Ethnic and Cultural Origins"
DATA_SHEET = "Discretionary Funding Requests"

TAXONOMY_ENTRY1 = "Ethnic and Cultural Origins Level 1"
TAXONOMY_ENTRY2 = "Ethnic and Cultural Origins Level 2"
TAXONOMY_ENTRY3 = "Ethnic and Cultural Origins Level 3"
TAXONOMY_ALL_TERMS = "All Terms"
TAXONOMY_SCOPE_NOTES = "Definitions and Scope Notes"

# Search priority order. All 4 are concatenated, but this order
# determines which match "wins" on ties (first column listed wins)
INPUT_COLS_PRIORITY = [
    "Final_Project_Description",
    "Final_Summary_Description",
    "Purpose",
    "Funding Request Name",
]

# Name/body split (Plan.md Phase 1, D1): a signal must be corroborated in the
# served-population body text to classify; a name-only signal degrades to
# General + flag. See get_body_and_name_texts().
BODY_COLS = ["Final_Project_Description", "Final_Summary_Description", "Purpose"]
NAME_COL  = "Funding Request Name"

OUTPUT_ETHNIC1 = "Ethnic 1 - FR6"
OUTPUT_ETHNIC2 = "Ethnic 2 - FR7"
OUTPUT_ETHNIC3 = "Ethnic 3 - FR8"
OUTPUT_FLAG = "Classification Flag"
OUTPUT_SEMANTIC = "Semantic Suggestion (REVIEW)"


# =======
# HELPERS
# =======

def clean_taxonomy_value(val):
    """Lowercase, strip, and remove the literal word 'origins'. Treats NaN/None as
    empty. Pandas reads blank Excel cells as float('nan'), so an explicit pd.isna() check is required here."""
    if val is None or (isinstance(val, float) and pd.isna(val)) or pd.isna(val):
        return ""
    val = str(val).strip()
    if val.lower() == "nan":
        return ""
    val = val.lower()
    val = val.replace("origins", "")
    return val.strip()

def safe_display(val):
    """Return the original display value for level1/level2/level3,
    treating NaN as empty string instead of the literal 'nan'."""
    if val is None or pd.isna(val):
        return ""
    val = str(val).strip()
    return "" if val.lower() == "nan" else val

def extract_match_keyword(raw_keyword):
    """Trim a taxonomy label down to the bare word(s) worth searching
    for in funding-request text: "black, not otherwise specified" ->
    "black", "indian (india)" -> "indian"."""
    keyword = raw_keyword.split(",")[0]
    keyword = re.sub(r"\s*\([^)]*\)\s*$", "", keyword)
    return keyword.strip()

def normalize_text(text):
    """Lowercase, strip punctuation, collapse whitespace. Used on the
    funding-request free text columns (not the taxonomy sheet, which
    uses clean_taxonomy_value instead)."""
    if pd.isna(text) or not isinstance(text, str):
        return ""
    text = text.lower()
    text = re.sub(r'[^\w\s]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

# UNCOMMENT : For implementation of alias check
# ------
# def apply_aliases(text):
#     for alias, canonical in KEYWORD_ALIASES.items():
#         text = re.sub(r'\b' + re.escape(alias) + r'\b', canonical, text)
#     return text

def matches_any(patterns, text):
    return any(re.search(p, text, re.IGNORECASE) for p in patterns)

def keyword_context_window(keyword, text, window=80):
    """Return the text immediately preceding the first occurrence of
    keyword, used to check for negation/example/override phrases that
    appear right before it."""
    pattern = re.compile(r'\b' + re.escape(keyword) + r'\b', re.IGNORECASE)
    m = pattern.search(text)
    if not m:
        return None
    start = max(0, m.start() - window)
    return text[start:m.start()]

def is_non_prefixed(keyword, text):
    """Return True if keyword appears immediately after 'non-' or 'non ' in text."""
    return bool(re.search(r'\bnon[\- ]' + re.escape(keyword), text, re.IGNORECASE))

def is_negated(keyword, text):
    snippet = keyword_context_window(keyword, text)
    return (
        (snippet is not None and matches_any(NEGATION_PHRASES, snippet))
        or is_non_prefixed(keyword, text)
    )

def is_example_mention(keyword, text):
    snippet = keyword_context_window(keyword, text)
    return snippet is not None and matches_any(EXAMPLE_PHRASES, snippet)

def phrase_has_serving_context(pattern, text, window=100):
    for m in re.finditer(pattern, text, re.IGNORECASE):
        start = max(0, m.start() - window)
        end = min(len(text), m.end() + window)
        if matches_any(SERVING_CONTEXT_WORDS, text[start:end]):
            return True
    return False

# ============================================================
# TAXONOMY BUILDER (Case 1-3 source data)
# Sort: deepest first, then LONGEST keyword first within same depth. Prevents "African"
# from matching before "Southern and East African".
# ============================================================

# ============================================================
# TAXONOMY BUILDER (Case 1-3 source data)
#
# Pulled from Column D "All Terms", a flat concatenation of
# Level 1 + Level 2 + Level 3 with the literal word "Origins" acting
# as the delimiter after L1 and L2 (no separator after L3,
# since specific terms like "Somali" don't carry an "Origins" suffix).
#
# Example: "African OriginsSouthern and East African OriginsSomali"
#   -> split on "Origins" -> ["African", "Southern and East African", "Somali"]
#
# Falls back to reading Level 1/2/3 columns directly if All Terms is
# blank or missing for a given row, for resilience against a broken
# formula cell.
#
# Sort: deepest first, then LONGEST keyword first within the same
# depth. This is the fix from the production skeleton — prevents
# "African" from matching before "Southern and East African".
# ============================================================

def parse_all_terms(all_terms_value):
    """
    Split the 'All Terms' cell on the literal word 'Origins' into
    its component level strings, stripped of whitespace. Trailing chars removed.
    """
    if pd.isna(all_terms_value) or not isinstance(all_terms_value, str):
        return []
    parts = all_terms_value.split("Origins")
    parts = [p.strip() for p in parts if p.strip()]
    return parts

def build_taxonomy(tax_df):
    entries = []
    for _, row in tax_df.iterrows():
        all_terms_raw = row.get(TAXONOMY_ALL_TERMS, "")
        parts = parse_all_terms(all_terms_raw)

        if parts:
            # Primary path: parsed from All Terms column.
            # The split strips the literal word "Origins" from every
            # fragment, but in the real taxonomy, Level 1 and Level 2
            # display values usually INCLUDE the "Origins" suffix
            # (e.g. "African Origins", "Southern and East African Origins")
            # -- except entries like "Black, not otherwise specified" and
            # "Jewish" under "Other Ethnic and Cultural Origins", which
            # only have "Origins" on Level 1. Count actual occurrences of
            # "Origins" in the source string so we know how many leading
            # parts really need it re-appended, instead of assuming L1+L2
            # always do.
            origins_count = all_terms_raw.count("Origins")
            display_parts = []
            for i, p in enumerate(parts):
                display_parts.append(f"{p} Origins" if i < origins_count else p)

            bare_l1 = parts[0] if len(parts) >= 1 else ""
            bare_l2 = parts[1] if len(parts) >= 2 else ""
            bare_l3 = parts[2] if len(parts) >= 3 else ""

            level1 = display_parts[0] if len(display_parts) >= 1 else ""
            level2 = display_parts[1] if len(display_parts) >= 2 else ""
            level3 = display_parts[2] if len(display_parts) >= 3 else ""

            depth   = len(parts)
            # Trim administrative qualifiers (", not otherwise specified",
            # "(India)") down to the bare searchable word -- funding
            # descriptions say "Black", never "Black, not otherwise specified".
            raw_keyword = (bare_l3 or bare_l2 or bare_l1).strip().lower()
            keyword = extract_match_keyword(raw_keyword)
        else:
            # Fallback: read Level 1/2/3 columns directly if 'All Terms'
            # is blank/missing for this row. These already include the
            # "Origins" suffix, so no re-appending needed.
            level1 = safe_display(row.get(TAXONOMY_ENTRY1, ""))
            level2 = safe_display(row.get(TAXONOMY_ENTRY2, ""))
            level3 = safe_display(row.get(TAXONOMY_ENTRY3, ""))
            if not level1:
                continue
            depth   = 3 if level3 else (2 if level2 else 1)
            raw_keyword = clean_taxonomy_value(level3 or level2 or level1)
            keyword = extract_match_keyword(raw_keyword)

        if not level1:
            continue

        entries.append({
            "keyword": keyword,
            "level1":  level1,
            "level2":  level2,
            "level3":  level3,
            "depth":   depth,
        })

    # Deepest first, then longest keyword first within the same depth
    entries.sort(key=lambda x: (-x["depth"], -len(x["keyword"])))
    return entries

# ==============================
# CASE-BY-CASE DETECTION LAYERS
# ==============================

def find_taxonomy_matches(text, taxonomy_entries):
    """
    Cases 1-3: direct taxonomy keyword matches, respecting negation
    and example-mention guards. Returns ALL valid matches found.
    The trailing s? allows plural forms ("Somalis", "Africans") to
    match the same singular taxonomy keyword without needing a
    separate entry for every pluralized variant.
    """
    matched = []
    for entry in taxonomy_entries:
        kw = entry["keyword"]
        if not kw:
            continue
        pattern = re.compile(r'\b' + re.escape(kw) + r's?\b', re.IGNORECASE)
        if not pattern.search(text):
            continue
        if is_negated(kw, text) or is_example_mention(kw, text):
            continue
        matched.append(entry)
    return matched

def find_pattern_match(text):
    """
    Case 4: structured/directional phrases not directly in taxonomy.
    Returns (l1, l2, l3, depth) or None.
    """
    for pattern, l1, l2, l3 in PATTERN_RULES:
        m = re.search(pattern, text, re.IGNORECASE)
        if not m:
            continue
        snippet = text[max(0, m.start()-80):m.start()]
        if matches_any(NEGATION_PHRASES, snippet):
            continue
        depth = 3 if l3 else (2 if l2 else 1)
        return (l1, l2, l3, depth)
    return None

def find_country_match(text):
    """
    Cases 6 & 7: nationality OR 'from <country>' phrasing.
    Returns (l1, l2, l3, depth) or None. The trailing s? allows plural
    forms ("Jamaicans", "Mexicans") to match without a separate entry
    for every pluralized variant.
    """
    for country, (l1, l2, l3) in COUNTRY_REGION_MAP.items():
        direct = re.search(r'\b' + re.escape(country) + r's?\b', text, re.IGNORECASE)
        from_phrase = re.search(r'\bfrom\s+' + re.escape(country) + r's?\b', text, re.IGNORECASE)
        m = direct or from_phrase
        if not m:
            continue
        snippet = text[max(0, m.start()-80):m.start()]
        if matches_any(NEGATION_PHRASES, snippet):
            continue
        depth = 3 if l3 else (2 if l2 else 1)
        return (l1, l2, l3, depth)
    return None

def is_bipoc_real_target(text):
    """Case 9 with context awareness: BIPOC mentioned as the actual
    target population vs. mentioned only as a general equity/mission
    statement (e.g. 'advancing equity', 'fostering BIPOC visibility'
    as one goal among several). We treat BIPOC as the target UNLESS
    it's wrapped in example-mention or override phrasing, or unless
    it appears alongside language suggesting it's an aspirational/
    values statement rather than a description of who is served.
    NOTE: This is ambiguous so: flagged
    for review rather than silently resolved either way.
    """
    found = False
    for kw_pattern in BIPOC_KEYWORDS:
        m = re.search(kw_pattern, text, re.IGNORECASE)
        if m:
            snippet = text[max(0, m.start()-80):m.start()]
            if matches_any(EXAMPLE_PHRASES, snippet):
                continue
            if matches_any(NEGATION_PHRASES, snippet):
                continue
            found = True
            break
    return found

def check_grassroots_case(text, has_ethnic_signal):
    """
    Case 11 (expanded): 'grassroots' / 'marginalized' / 'ethnocultural' /
    'multicultural' / 'refugee' / 'immigrant' etc. never count as a signal
    on their own. Unlike negation/example-mention, this never suppresses
    silently either way -- caller flags the result whether a real signal
    is present alongside the buzzword or not.

    has_ethnic_signal is computed by the CALLER from the same candidate
    pool used for final classification in classify_row() -- this used
    to re-derive its own signal-detection pass via a separate
    has_real_ethnic_signal() helper, which silently missed BIPOC/POC
    (and could drift out of sync again the next time a new detection
    layer gets added to one path but not the other). Taking the
    already-computed signal as a parameter makes that class of bug
    structurally impossible -- there is only one detection pass now.
    """
    has_ambiguous = matches_any(AMBIGUOUS_EQUITY_WORDS, text)
    if not has_ambiguous:
        return None  # not relevant, no opinion
    if has_ethnic_signal:
        return "has_signal"  # real signal present -- classify normally, caller still flags
    return "no_signal"  # ambiguous word present but no real ethnic keyword

def check_org_name_lookup(text):
    """
    Case 10: known organization names. Only called as last resort.
    """
    for org_name, (l1, l2, l3) in ORG_NAME_ETHNICITY_MAP.items():
        if re.search(r'\b' + re.escape(org_name) + r'\b', text, re.IGNORECASE):
            return (l1, l2, l3)
    return None

def detect_broad_identity(text):
    return matches_any(BROAD_IDENTITY_KEYWORDS, text)

def context_is_overridden(text):
    return any(phrase_has_serving_context(p, text) for p in EXPANSION_PHRASES)

def context_is_historical(text):
    return any(phrase_has_serving_context(p, text) for p in HISTORICAL_PHRASES)

def context_is_aspirational(text):
    return matches_any(ASPIRATIONAL_PHRASES, text)

# =================================================
# CONTEXT SIGNAL LAYER
# Detects soft discourse signals in the full combined text.
# Results are annotations only — they MUST NOT affect classification decisions.
# =================================================

def extract_context_signals(text):
    return {
        "historical":   matches_any(HISTORICAL_PHRASES, text),
        "expansion":    matches_any(EXPANSION_PHRASES, text),
        "aspirational": matches_any(ASPIRATIONAL_PHRASES, text),
        "negation":     matches_any(ETHNIC_ANNOTATION_NEGATION_PHRASES, text),
        "example":      matches_any(EXAMPLE_PHRASES, text),
    }

def build_context_notes(signals, ethnic_term_present=False):
    """
    Production annotation notes.

    Only negation is surfaced — it directly affects how confident a reviewer
    should be in the ethnic evidence found (a negated term may or may not
    indicate the population served).

    Historical, expansion, aspirational, and example signals are omitted here
    because they are extremely common in nonprofit funding language and create
    alert fatigue when surfaced on every request.  They remain detectable via
    extract_context_signals() and are available in full via
    build_debug_context_notes() for debug overlay use.
    """
    notes = []
    # Anchor: only surface negation when an ethnic term was actually detected.
    if signals["negation"] and ethnic_term_present:
        notes.append("Negation detected - verify exclusion vs inclusion intent")
    return notes

def build_debug_context_notes(signals):
    """
    Full context annotation notes including discourse-framing signals.
    For debug overlay use only — NOT emitted in production output.
    """
    notes = []
    if signals["historical"]:
        notes.append("Historical framing detected - may refer to past service scope only")
    if signals["expansion"]:
        notes.append("Scope expansion language detected - indicates broadened or non-exclusive targeting")
    if signals["aspirational"]:
        notes.append("Aspirational/future-oriented language detected - may not reflect current service population")
    if signals["negation"]:
        notes.append("Negation detected - verify exclusion vs inclusion intent")
    if signals["example"]:
        notes.append("Example-based phrasing detected - referenced group may not be primary target")
    return notes

# =================================================
# CASE 13 — Potential Ethnocultural Organization Name
#
# Safety-net detector for unknown ethnocultural org names in the
# Funding Request Name column.  Only fires when Engine 1 produced zero
# candidates and BIPOC is absent — i.e. the row would otherwise fall
# through to General Population with no ethnic signal at all.
#
# NEVER classifies. NEVER overrides. Review flag only.
# =================================================

def looks_like_demonym(group_name):
    """Return True if group_name could plausibly be an ethnic/national identifier.
    Primary gate: any token in NON_ETHNIC_LEADING_WORDS → almost certainly not ethnic."""
    tokens = group_name.lower().split()
    return not any(t in NON_ETHNIC_LEADING_WORDS for t in tokens)

def is_known_taxonomy_keyword(group_name, taxonomy_entries):
    """Return True if group_name contains a recognized taxonomy keyword."""
    for entry in taxonomy_entries:
        kw = entry.get("keyword", "")
        if not kw:
            continue
        if re.search(r'\b' + re.escape(kw) + r's?\b', group_name, re.IGNORECASE):
            return True
    return False

def matches_pattern_rule(group_name):
    """Return True if group_name matches any PATTERN_RULES entry."""
    for pattern, *_ in PATTERN_RULES:
        if re.search(pattern, group_name, re.IGNORECASE):
            return True
    return False

def detect_ethnocultural_org_name(funding_request_name, candidates, bipoc_present, taxonomy_entries):
    """
    Case 13 — low-recall, high-precision safety net.

    Detects potential ethnocultural organization names in the Funding Request
    Name that Engine 1 did not already classify.

    Guard clause (Step 1): if ANY Engine 1 candidate exists or BIPOC is
    present, the row is already being handled — return None immediately.
    This ensures Case 13 only triggers as a true last resort.

    Negative filters (Step 6): even when the title matches the org-name
    pattern, do NOT flag if the extracted group token is:
      - a recognized taxonomy keyword
      - a known country/region/nationality in COUNTRY_REGION_MAP
      - matched by any PATTERN_RULES entry (directional ethnonyms etc.)
    """
    if candidates or bipoc_present:
        return None

    title = normalize_text(funding_request_name or "")
    if not title:
        return None

    for pattern in CASE_13_PATTERNS:
        match = re.search(pattern, title, re.IGNORECASE)
        if not match:
            continue

        group_name = match.group(1).strip()

        if is_known_taxonomy_keyword(group_name, taxonomy_entries):
            return None
        if group_name in COUNTRY_REGION_MAP:
            return None
        if matches_pattern_rule(group_name):
            return None
        if not looks_like_demonym(group_name):
            return None

        return "Note (low priority): Potential ethnocultural organization name detected - verify group identity manually"

    return None

# =================================================
# TEXT EXTRACTION — concatenate across all 4 columns
# =================================================

# Some compound identity phrases mean something specific that's
# different from their literal sub-words. "African Canadian" names the
# Black Canadian identity (-> Other Ethnic and Cultural Origins / Black,
# not otherwise specified) -- it is NOT the same as "African" alone
# (-> African Origins), but the bare word "african" matches that
# taxonomy entry on its own regardless of what follows it, so the
# phrase was getting silently reduced to the wrong, broader category.
# Rewriting the phrase to the word that already has a correct taxonomy
# entry means it flows through the EXISTING matching pipeline (negation
# guards, multi-group detection, depth resolution) automatically --
# no separate override system needed, and a second distinct group
# mentioned elsewhere in the same text (e.g. "African Canadian and
# Somali") still correctly produces Multiple instead of silently
# picking one and ignoring the other.
def apply_identity_phrase_rewrites(text):
    for pattern, replacement in IDENTITY_PHRASE_REWRITES:
        def _replace(m, text=text):
            preceding = text[max(0, m.start()-12):m.start()].strip(" -")
            if any(preceding.endswith(prefix) for prefix in DIRECTIONAL_AFRICAN_PREFIXES):
                return m.group(0)  # leave a regional phrase like "East African Canadian" untouched
            return replacement
        text = re.sub(pattern, _replace, text, flags=re.IGNORECASE)
    return text

def get_column_texts(row):
    texts = []
    for col in INPUT_COLS_PRIORITY:
        val = row.get(col, "")
        raw = normalize_text(val)
        raw = apply_identity_phrase_rewrites(raw)
        # Commented out for alias check implementation.
        #raw = apply_aliases(raw)
        texts.append(raw)
    return texts

def get_body_and_name_texts(row):
    """Return (body_text, name_text), each normalized + identity-rewritten
    exactly like get_column_texts, so callers can require a served-population
    signal in the body and treat a name-only signal as unclassified."""
    def _prep(val):
        return apply_identity_phrase_rewrites(normalize_text(val))
    body = " ".join(t for t in (_prep(row.get(c, "")) for c in BODY_COLS) if t.strip())
    name = _prep(row.get(NAME_COL, ""))
    return body, name

# ==========================
# MAIN CLASSIFICATION LOGIC
# =============================

def classify_row(row, taxonomy_entries):
    col_texts = get_column_texts(row)
    combined  = " ".join(t for t in col_texts if t.strip())

    # Extra notes (buzzword / cultural association mentions) appended to whatever flag this function ends up returning, no matter which branch below resolves the classification.
    extra_notes = []

    def finalize(e1, e2, e3, flag):
        if extra_notes:
            note_text = "; ".join(extra_notes)
            flag = f"{flag}; {note_text}" if flag else note_text
        return (e1, e2, e3, flag)

    if not combined.strip():
        return finalize(GENERAL_POP, "", "", "Empty input")

    # Context signal layer — annotation only, no control flow.
    # All discourse signals (historical, expansion, aspirational, negation, example)
    # are recorded as flags. Classification is never branched on them.
    context_signals = extract_context_signals(combined)
    extra_notes = []

    # "Cultural Association" often hides a specific named group (e.g.
    # "Kerala Cultural Association") -- always worth a second look,
    # regardless of how this row otherwise resolves.
    if re.search(r"\bcultural association\b", combined, re.IGNORECASE): # Add check for {group} cultur(e/al) or association
        extra_notes.append("'Cultural Association' detected - verify named group manually")

    # Cases 1-3: direct taxonomy matches, across FR columns
    all_matches = []
    for col_text in col_texts:
        if col_text.strip():
            all_matches.extend(find_taxonomy_matches(col_text, taxonomy_entries))

    seen = set()
    unique_matches = []
    for m in all_matches:
        if m["keyword"] not in seen:
            seen.add(m["keyword"])
            unique_matches.append(m)

    # Build a unified candidate pool: taxonomy + compound + pattern +
    # country + broad identity. Each candidate: (level1, level2, level3, depth, source)
    # Built BEFORE the buzzword check below so that check has the SAME
    # evidence the rest of this function uses -- see check_grassroots_case().
    candidates = []
    for m in unique_matches:
        candidates.append((m["level1"], m["level2"] or "", m["level3"] or "",
                            m["depth"], "taxonomy"))

    # Afro-Caribbean / Afro-Latino -- add BOTH halves as candidates so
    # they combine into Multiple the same way two separate mentions would.
    for compound_pattern, group_tuples in ALWAYS_MULTIPLE_COMPOUNDS.items():
        if re.search(compound_pattern, combined, re.IGNORECASE):
            for (l1, l2, l3) in group_tuples:
                candidates.append((l1, l2, l3, 1, "compound"))

    pattern_result = find_pattern_match(combined)
    if pattern_result:
        l1, l2, l3, depth = pattern_result
        candidates.append((l1, l2, l3, depth, "pattern"))

    country_result = find_country_match(combined)
    if country_result:
        l1, l2, l3, depth = country_result
        candidates.append((l1, l2, l3, depth, "country"))

    # Broad identity labels with no specific taxonomy entry (Hispanic,
    # Latino, etc. -- NOT Black/Jewish/Arab, real taxonomy entries now,
    # already came through find_taxonomy_matches above). Folded into
    # the same pool so they can combine into Multiple alongside a
    # separately-named group too.
    if detect_broad_identity(combined):
        candidates.append((OTHER_ETHNIC, "", "", 1, "broad_identity"))

    # Dedupe by actual (level1, level2, level3) outcome; two different
    # detection methods (e.g. pattern rule + country map) landing on the
    # exact same conclusion is a single confirmed answer, not a multi-
    # group signal. Keeps the first-seen source label for the flag text.
    seen_outcomes = set()
    deduped_candidates = []
    for c in candidates:
        outcome_key = (c[0], c[1], c[2])
        if outcome_key not in seen_outcomes:
            seen_outcomes.add(outcome_key)
            deduped_candidates.append(c)
    candidates = deduped_candidates

    # Case 9: BIPOC (context-aware), computed here -- before the buzzword
    # check below -- so it can feed into has_ethnic_signal too.
    bipoc_present = is_bipoc_real_target(combined)

    # Case 11: grassroots / ambiguous equity words -- never suppress
    # silently now, always flagged whether or not a real signal is present.
    # has_ethnic_signal reads straight from the candidate pool and
    # bipoc_present above -- the exact same evidence the rest of this
    # function uses, by construction. See check_grassroots_case().
    has_ethnic_signal = bool(candidates) or bipoc_present
    if context_signals["negation"] and has_ethnic_signal:
        extra_notes.insert(0, "Negation detected - verify exclusion vs inclusion intent")
    grassroots_state = check_grassroots_case(combined, has_ethnic_signal) # Handle 'ethnocultural', 'marginalized' as well, since they have the same rule as 'grassroots'
    if grassroots_state == "no_signal":
        extra_notes.append("Note (low priority): Ambiguous equity term with no paired ethnic signal")
    if grassroots_state == "has_signal":
        extra_notes.append("Note (low priority): Equity/diversity buzzword present alongside a real signal - verify manually")

    # Possible consulted-party mention (expert/advisor/biologist role)
    # rather than the population served -- flagged, never suppressed.
    if candidates and matches_any(EXPERT_ROLE_PHRASES, combined):
        extra_notes.append("Note (low priority): Possible consulted-party mention (expert/advisor role) rather than served population - verify manually")

    # Checked here (after candidates are gathered) so we can tell whether
    # BIPOC is mentioned ALONGSIDE a real specific group (the ambiguous nuance from the README) vs. BIPOC being the only signal.
    if bipoc_present:
        if candidates:
            # BIPOC + a specific named group both present. Do not silently resolve, flag for manual review. # We can set this as a generic 'Ambiguous: BIPOC mentioned alongside specific group(s) flag so number of unique classification flags are smaller
            other_groups = sorted(set(c[0] for c in candidates))
            flag = ("Ambiguous: BIPOC mentioned alongside specific group(s) ("
                    + ", ".join(other_groups) + ") - verify manually")
        else:
            flag = "BIPOC signal detected"
        return finalize(MULTIPLE_ETHNIC, "", "", flag)

    # Resolve candidates: distinct Level 1 groups ANYWHERE in the pool
    # (not just tied at the deepest level) mean Multiple -- lets "African" (depth 1) combine with "Black" (depth 2, a different L1 branch
    # under "Other Ethnic and Cultural Origins") instead of the deeper one silently winning by depth alone.
    if candidates:
        distinct_l1 = set(c[0] for c in candidates)

        if len(distinct_l1) >= 2:
            return finalize(MULTIPLE_ETHNIC, "", "", "Review: multiple distinct groups detected")

        # All candidates share one Level 1 -- use depth to pick the most
        # specific. If several DIFFERENT branches still tie at the
        # deepest level within this one Level 1, that's still Multiple.
        max_depth = max(c[3] for c in candidates)
        deepest = [c for c in candidates if c[3] == max_depth]

        if len(deepest) >= 2:
            return finalize(MULTIPLE_ETHNIC, "", "", "Review: multiple sub-groups within same origin")

        l1, l2, l3, depth, source = deepest[0]
        flag = ""
        if source == "pattern":
            flag = "Pattern rule match (Directional phrase, e.g. North African)"
        elif source == "country":
            flag = "Country/nationality mapping match"
        elif source == "compound":
            flag = "Compound identity term match"
        elif source == "broad_identity":
            flag = "Broad identity term - review recommended"
        return finalize(l1, l2, l3, flag)

    # Case 10: organization name lookup (LAST RESORT before General) NOTE: Need to highlight Black Canadian Women to flag
    org_result = check_org_name_lookup(combined)
    if org_result:
        l1, l2, l3 = org_result
        return finalize(l1, l2, l3, "Matched via known organization name lookup")

    # Case 12: fallback
    return finalize(GENERAL_POP, "", "", "")

# =====
# MAIN
# =====
def main():
    # Deferred import avoids circular dependency: classify_pipeline imports from this module.
    from classify_pipeline import classify_row as pipeline_classify_row

    start_time = time.time()

    taxonomy_filepath = bootstrap.PROJECT_ROOT / "Taxonomy" / "Taxonomy - Definitions.xlsx"
    funding_filepath = bootstrap.PROJECT_ROOT / "Data Sheets" / "FR testing.xlsx"
 
    print(f"Loading taxonomy from: {taxonomy_filepath}")
    try:
        tax_df = pd.read_excel(taxonomy_filepath, sheet_name=TAXONOMY_SHEET, dtype=str)
    except Exception as e:
        print(f"Error loading taxonomy sheet '{TAXONOMY_SHEET}' from '{taxonomy_filepath}': {e}")
        sys.exit(1)
 
    print(f"Loading funding requests from: {funding_filepath}")
    try:
        data_df = pd.read_excel(funding_filepath, sheet_name=DATA_SHEET, dtype=str)
    except Exception as e:
        print(f"Error loading data sheet '{DATA_SHEET}' from '{funding_filepath}': {e}")
        sys.exit(1)
 
    print(f"Taxonomy rows: {len(tax_df)} | Data rows: {len(data_df)}")
 
    taxonomy_entries = build_taxonomy(tax_df)
    print(f"Taxonomy entries parsed: {len(taxonomy_entries)}")

    # Semantic Fallback - Level 2 Engine
    semantic_entries = None
    semantic_embeddings = None

    if SEMANTIC_AVAILABLE:
        scope_notes = semantic_fallback.build_scope_note_map(
            tax_df, TAXONOMY_ENTRY1, TAXONOMY_SCOPE_NOTES, safe_display)
        semantic_entries, semantic_texts = semantic_fallback.build_candidate_texts(
            taxonomy_entries, scope_notes)
        semantic_embeddings = semantic_fallback.get_taxonomy_embeddings(semantic_texts)
    else:
        print("semantic_fallback not available (sentence-transformers not installed) — skipping semantic suggestions.")
 
    for col in [OUTPUT_ETHNIC1, OUTPUT_ETHNIC2, OUTPUT_ETHNIC3, OUTPUT_FLAG, OUTPUT_SEMANTIC]:
        if col not in data_df.columns:
            data_df[col] = ""
    # Count how many rows fall into each outcome bucket, for summary stats at the end. Note that these are not mutually exclusive categories (e.g. a row with a pattern match that's also flagged for aspirational language would count in both "pattern" and "flagged"), but they give a general sense of how many hits came from each detection method and how many were flagged for review.
    stats = {"3-level": 0, "2-level": 0, "1-level": 0, "multiple": 0,
              "other": 0, "general": 0, "flagged": 0, "pattern": 0,
              "country": 0, "org_lookup": 0, "grassroots_filtered": 0, "semantic_suggested": 0}
 
    for idx, row in data_df.iterrows():
        # Route through the canonical refactored pipeline (classify_pipeline → resolver)
        e1, e2, e3, flag = pipeline_classify_row(row, taxonomy_entries)
        data_df.at[idx, OUTPUT_ETHNIC1] = e1
        data_df.at[idx, OUTPUT_ETHNIC2] = e2
        data_df.at[idx, OUTPUT_ETHNIC3] = e3
        data_df.at[idx, OUTPUT_FLAG] = flag

        # Engine 2 - This is our Semantic fallback layer, which only triggers if Engine 1 returns General Population (i.e. no
        if SEMANTIC_AVAILABLE and e1 == GENERAL_POP:
            combined_text = " ".join(t for t in get_column_texts(row) if t.strip())
            suggestion = semantic_fallback.find_semantic_suggestion(
                combined_text, semantic_entries, semantic_embeddings)
            if suggestion:
                sl1, sl2, sl3, score, margin = suggestion
                parts = [p for p in [sl1, sl2, sl3] if p]
                data_df.at[idx, OUTPUT_SEMANTIC] = f"{' / '.join(parts)} (similarity: {score:.2f}, margin: {margin:.2f})"
                stats["semantic_suggested"] += 1
 
        if e1 == MULTIPLE_ETHNIC:
            stats["multiple"] += 1
        elif e1 == OTHER_ETHNIC:
            stats["other"] += 1
        elif e1 == GENERAL_POP:
            stats["general"] += 1
        elif e3:
            stats["3-level"] += 1
        elif e2:
            stats["2-level"] += 1
        else:
            stats["1-level"] += 1
 
        if "pattern rule" in flag.lower():
            stats["pattern"] += 1
        if "country" in flag.lower():
            stats["country"] += 1
        if "organization name" in flag.lower():
            stats["org_lookup"] += 1
        if "ambiguous equity" in flag.lower():
            stats["grassroots_filtered"] += 1
        if flag:
            stats["flagged"] += 1
 
    wb = load_workbook(funding_filepath)
    ws = wb[DATA_SHEET]
    headers = {cell.value: cell.column for cell in ws[1]}
 
    for col_name in [OUTPUT_ETHNIC1, OUTPUT_ETHNIC2, OUTPUT_ETHNIC3, OUTPUT_FLAG, OUTPUT_SEMANTIC]:
        if col_name not in headers:
            new_col = ws.max_column + 1
            ws.cell(row=1, column=new_col, value=col_name)
            headers[col_name] = new_col
 
    for i, (idx, row) in enumerate(data_df.iterrows(), start=2):
        ws.cell(row=i, column=headers[OUTPUT_ETHNIC1], value=data_df.at[idx, OUTPUT_ETHNIC1])
        ws.cell(row=i, column=headers[OUTPUT_ETHNIC2], value=data_df.at[idx, OUTPUT_ETHNIC2])
        ws.cell(row=i, column=headers[OUTPUT_ETHNIC3], value=data_df.at[idx, OUTPUT_ETHNIC3])
        ws.cell(row=i, column=headers[OUTPUT_FLAG],    value=data_df.at[idx, OUTPUT_FLAG])
        ws.cell(row=i, column=headers[OUTPUT_SEMANTIC], value=data_df.at[idx, OUTPUT_SEMANTIC]) # Writes "" for every row that didn't receive a suggestion (General pop.)
    wb.save(funding_filepath)
    
    print("\nResults:")
    for k, v in stats.items():
        print(f"  {k}: {v}")

    print(f"\nOutput written to: {funding_filepath}")
    elapsed = time.time() - start_time
    print(f"Ethnic classification completed in {elapsed:.1f} seconds.")
 
if __name__ == "__main__":
    main()
