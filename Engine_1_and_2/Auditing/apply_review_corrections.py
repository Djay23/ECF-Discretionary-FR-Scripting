import sys
import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))  # Engine 1 and 2
import bootstrap  # noqa: F401  (imported for effect: sys.path + UTF-8 stdout,
                  # which the em-dashes in this script's output depend on)
from dataset_config import DATASETS

"""
apply_review_corrections.py
---------------------------
Moves the corrections you made in the review workbook into the FINAL REVIEW
copies (never the authoritative gold).

Flow:
    1. Manual correction of Classification Review.xlsx (the mirror of the
       gold, one sheet per dataset).
    2. This script diffs each review sheet against the matching Final Review
       copy, matched by SF_18_ID_Funding_Request__c (never row position), and
       reports every cell you changed in the classification/audit columns.
    3. With --apply, it writes ONLY those changed cells into the Final Review
       copy. Source text, IDs, and everything outside SYNC_COLS are untouched.

Default is a DRY RUN: it prints every pending change and writes nothing.
Add --apply to actually write to the Final Review copies.

    python Engine_1_and_2/Auditing/apply_review_corrections.py           # dry run
    python Engine_1_and_2/Auditing/apply_review_corrections.py --apply   # write
"""

import paths

REVIEW_FILE = paths.REVIEW_FILE

# Final Review copies to write to (NOT the authoritative gold). Discovered by
# matching the dataset year in the filename -- renaming a workbook no longer
# requires editing this file.
POST_REVIEW = {name: d.final_review
               for name, d in paths.discover().items()
               if d.final_review}

DATA_SHEET = paths.DATA_SHEET
ID_COL = "SF_18_ID_Funding_Request__c"

# Only these columns are ever compared/written. Source text, IDs, and the
# evidence column are deliberately excluded.
SYNC_COLS = [
    "Ethnic 1 - FR6", "Ethnic 2 - FR7", "Ethnic 3 - FR8",
    "Gender Id - FR9", "Sexual Id - FR10",
    "Sector 1 - FR2", "Sector 2 - FR3", "Sector 3 - FR4", "Sector 4 - FR5",
    "ECD - FR", "AH - FR",
    "Classification Flag", "Gender Classification Flag", "Sexual Classification Flag",
    "Classification Accuracy (Corrected in Different Areas)",
]


def _read_safe(path, **kwargs):
    """Read that survives an Excel/OneDrive lock via a temp copy."""
    try:
        return pd.read_excel(path, **kwargs)
    except PermissionError:
        import shutil, tempfile
        tmp = Path(tempfile.gettempdir()) / f"_locked_{Path(path).name}"
        shutil.copy2(path, tmp)
        try:
            return pd.read_excel(tmp, **kwargs)
        finally:
            try:
                tmp.unlink()
            except OSError:
                pass


def _s(v):
    return "" if v is None else str(v).strip()


class ReviewSyncError(Exception):
    """A setup problem that must stop the run (missing file, missing sheet).
    Reported as a plain message, never a traceback."""


def diff_dataset(name):
    """Return (list of changes, review_sheet_name) for one dataset. A change is
    (id, request_name, column, old, new). Raises ReviewSyncError if the review
    sheet or the Final Review copy is missing — a missing target is a setup
    error, never a silent skip."""
    sheet = name.replace("_", "-")
    pr_path = POST_REVIEW.get(name)
    if pr_path is None:
        raise ReviewSyncError(
            f"[{name}] no Final Review copy found for dataset '{name}'.\n"
            f"           Put one in: {paths.FINAL_REVIEW_DIR}\n"
            f"           Its filename must contain the same year, e.g. "
            f"'Discretionary Funding Requests - {name} (GOLD) Final review.xlsx'.")
    if not pr_path.exists():
        raise ReviewSyncError(
            f"[{name}] Final Review copy has gone missing since it was found:\n"
            f"           {pr_path}\n"
            f"           It was probably renamed, moved, or is still syncing.")
    try:
        review = _read_safe(REVIEW_FILE, sheet_name=sheet, dtype=str).fillna("")
    except ValueError:
        raise ReviewSyncError(
            f"[{name}] review sheet '{sheet}' not found in {REVIEW_FILE.name}. "
            f"Each dataset needs a sheet named after it with '_' as '-'.")
    pr = _read_safe(pr_path, sheet_name=DATA_SHEET, dtype=str).fillna("")

    pr_by_id = {_s(r[ID_COL]): r for _, r in pr.iterrows()}
    cols = [c for c in SYNC_COLS if c in review.columns and c in pr.columns]

    changes = []
    for _, rrow in review.iterrows():
        rid = _s(rrow[ID_COL])
        prow = pr_by_id.get(rid)
        if prow is None:
            continue
        for col in cols:
            old, new = _s(prow[col]), _s(rrow[col])
            if old != new:
                changes.append((rid, _s(rrow.get("Funding Request Name", ""))[:45], col, old, new))
    return changes, sheet


def apply_changes(name, changes):
    """Write the changed cells into the Final Review copy (matched by ID)."""
    pr_path = POST_REVIEW[name]
    wb = load_workbook(pr_path)
    ws = wb[DATA_SHEET]
    headers = {cell.value: cell.column for cell in ws[1]}
    id_to_row = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, headers[ID_COL]).value
        if v is not None:
            id_to_row[str(v).strip()] = r
    written = 0
    for rid, _nm, col, _old, new in changes:
        er = id_to_row.get(rid)
        if er and col in headers:
            ws.cell(er, headers[col], value=new)
            written += 1
    wb.save(pr_path)
    return written


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true",
                    help="write the changes into the Final Review copies (default: dry run)")
    args = ap.parse_args()

    if not REVIEW_FILE.exists():
        print(f"ERROR: review file not found:\n           {REVIEW_FILE}")
        return 1

    # Phase 1 — diff every dataset before writing anything, so a missing file or
    # sheet stops the whole run instead of silently syncing only some datasets.
    diffs, problems = {}, []
    for name in DATASETS:
        try:
            diffs[name] = diff_dataset(name)
        except ReviewSyncError as e:
            problems.append(str(e))

    if problems:
        print(f"\nABORTED — {len(problems)} dataset(s) could not be synced. "
              f"Nothing was written.\n")
        for p in problems:
            print(f"  ERROR: {p}")
        print(f"\nFix the above and re-run. Datasets that were fine "
              f"({', '.join(n for n in diffs) or 'none'}) were left untouched.")
        return 1

    # Phase 2 — everything resolved; report and (optionally) write.
    total = 0
    for name in DATASETS:
        changes, sheet = diffs[name]
        print(f"\n=== {name} (review sheet '{sheet}' -> {POST_REVIEW[name].name}) ===")
        print(f"  {len(changes)} cell change(s):")
        for rid, nm, col, old, new in changes:
            print(f"    {nm:47} {col:52} {old[:24]!r} -> {new[:24]!r}")
        total += len(changes)
        if args.apply and changes:
            try:
                n = apply_changes(name, changes)
            except PermissionError:
                print(f"  ERROR: {POST_REVIEW[name].name} is locked (open in Excel?). "
                      f"Close it and re-run — nothing was written for {name}.")
                return 1
            print(f"  applied {n} change(s) to {POST_REVIEW[name].name}")

    if not args.apply:
        print(f"\nDRY RUN — {total} change(s) pending. Re-run with --apply to write them "
              f"to the Final Review copies.")
    else:
        print(f"\nDONE — {total} change(s) written to the Final Review copies "
              f"(authoritative gold untouched).")
    return 0


if __name__ == "__main__":
    sys.exit(main())
