import sys
import os
import shutil
import argparse
import tempfile
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))  # Engine 1 and 2
import bootstrap
import ethnic_taggerv3 as et

"""
build_flag_review_column.py
---------------------------
Decision-support tool for the flag de-noising pass. Adds ONE new column,
"Flag Review (keep/drop/merge)", to FR testing.xlsx immediately after the
Sexual Classification Flag column, so the stakeholder can decide which review
flags earn their keep -- side by side with the real funding-request rows and
the existing ethnic/gender/sexual flag columns.

It NEVER changes a classification, edits an existing column, or removes an
engine flag. It only annotates the flags ALREADY present in the three flag
columns with a recommended disposition (KEEP / DROP) + a one-line rationale,
drawn from the static FLAG_DISPOSITIONS catalog below. Dispositions are my
recommendations only; the stakeholder overrides any cell by editing it.

Run:
    # dry run -- no write; prints disposition tally, catalog-coverage report,
    # and a drift check vs the live engine:
    python Engine_1_and_2/Auditing/build_flag_review_column.py --dry-run

    # real run -- backs up the workbook, then inserts the column in place.
    # FR testing.xlsx MUST be CLOSED in Excel first (exclusive-lock otherwise).
    python Engine_1_and_2/Auditing/build_flag_review_column.py
"""

ETHNIC_FLAG_COL = "Classification Flag"
GENDER_FLAG_COL = "Gender Classification Flag"
SEXUAL_FLAG_COL = "Sexual Classification Flag"
REVIEW_COL_HEADER = "Flag Review (keep/drop/merge)"

# ---------------------------------------------------------------------------
# FLAG DISPOSITIONS CATALOG (my keep/drop recommendations)
#
# Each rule: (kind, key, short_name, disposition, rationale)
#   kind "prefix"   -> normalized flag startswith key
#   kind "contains" -> key is a substring of the normalized flag
#   kind "all"      -> key is a list; every substring must be present
# Rules are tried in order; first match wins. Keys are lowercase; normalize()
# lowercases the flag and folds em/en dashes to '-' and collapses whitespace.
#
# Grounded in the locked "does it change a reviewer's action?" principle
# (memory: classification-policy-decisions). Flags that prevent a real
# misclassification, name a genuine ambiguity, or are policy-mandated
# transparency notes -> KEEP. Emphasis / provenance / high-volume redundant
# notes -> DROP (alert fatigue). Stakeholder overrides any of these.
# ---------------------------------------------------------------------------

FLAG_DISPOSITIONS = [
    # --- specific rules first ---------------------------------------------
    ("contains", "ethnic/cultural signal co-present", "Cross-domain co-signal",
     "DROP", "highest-volume note; intersectionality already shown in labels + multi_identity_markers"),
    ("all", ["especially", "particularly", "broader"], "Especially/particularly emphasis",
     "DROP", "rarely changes the label; classic alert-fatigue emphasis note"),
    ("contains", "'cultural association' detected", "Cultural Association detected",
     "DROP", "vague prompt; low action yield"),
    ("contains", "aspirational/future language", "Aspirational/future language",
     "DROP", "locked policy explicitly calls aspirational a noise flag"),

    # --- KEEP: prevents a real misclassification --------------------------
    ("contains", "francophone/language organization name only", "Francophone/language org name only",
     "KEEP", "same family as religious-only; prevents language->ethnicity error"),
    ("contains", "french reference treated as language accommodation", "French = language accommodation",
     "KEEP", "prevents French-language->ethnicity error"),
    ("contains", "'hindu' detected", "Hindu religion vs ethnicity",
     "KEEP", "prevents religion->ethnicity conflation; rare but correct"),

    # --- KEEP: name-only classification must be verified -------------------
    ("contains", "classified from organization name (silent body)", "Classified from org name (silent body)",
     "KEEP", "name-only classification must be verified (locked org-name policy)"),
    ("contains", "classified from the organization's name via a curated lookup", "Classified from curated org lookup",
     "KEEP", "name-only; verify served population"),
    ("contains", "signal appears only in the organization", "Signal only in org/request name",
     "KEEP", "name-only; verify served population"),
    ("contains", "appears inside an organization or proper name", "Gender/sexual term inside org name",
     "KEEP", "name-vs-served disambiguation"),
    ("contains", "potential ethnocultural organization name detected", "Potential ethnocultural org name",
     "KEEP", "high-precision unknown-org safety net"),

    # --- KEEP: locked policy signals --------------------------------------
    ("contains", "appears near a negation word", "Near a negation word",
     "KEEP", "negation is the one always-surfaced context signal (locked)"),
    ("contains", "black and indigenous co-present", "BIPOC (Black+Indigenous co-present)",
     "KEEP", "locked BIPOC policy; genuine ambiguity"),
    ("contains", "bipoc keyword alongside specific group", "BIPOC keyword alongside specific group(s)",
     "KEEP", "flags BIPOC+specific tension"),
    ("contains", "indigenous umbrella term co-occurs", "Indigenous umbrella + sub-group(s)",
     "KEEP", "genuinely ambiguous which sub-group is served"),
    ("prefix", "general indigenous term matched", "General Indigenous term, no nation named",
     "KEEP", "signals umbrella-level result a reviewer can narrow; actionable"),
    ("contains", "indigenous topic/partnership mention", "Indigenous topic/partnership mention",
     "KEEP", "targets the known Indigenous over-fire backlog; high value"),
    ("contains", "2slgbtqia+ inferred from a gender-identity term", "2SLGBTQIA+ inferred from gender term",
     "KEEP", "prevents orientation over-inference"),
    ("contains", "2slgbtqia+ umbrella acronym", "2SLGBTQIA+ umbrella acronym inferred",
     "KEEP", "prevents over-inference from the umbrella acronym"),
    ("contains", "ambiguous coded gender/orientation term", "Ambiguous coded gender/orientation term",
     "KEEP", "cross-axis verify (femme/masc/butch)"),

    # --- KEEP: locked transparency mandate ("why was X ignored?") ---------
    ("all", ["mentioned in", "context only"], "Group mentioned in <role> context only",
     "KEEP", "locked transparency mandate; explains why a mention was not classified"),
    ("contains", "mentioned via org-name self-reference in body", "Mentioned via org-name self-reference",
     "KEEP", "transparency note; org self-reference, not served"),
    ("contains", "incidental family context", "Incidental family context",
     "KEEP", "transparency; targets family-context backlog"),
    ("contains", "example/illustrative", "Example/illustrative mention",
     "KEEP", "transparency note; illustrative mention, not classified"),

    # --- KEEP: Multiple composition (put last; broad prefix) --------------
    ("prefix", "multiple:", "Multiple composition",
     "KEEP", "shows which identities composed Multiple; actionable"),
]


def normalize(s):
    s = str(s).lower().replace("—", "-").replace("–", "-")
    return " ".join(s.split())


def classify_flag(atomic):
    """Return (short_name, disposition, rationale) for one atomic flag, or
    None if no catalog rule matches (surfaced as UNCATALOGUED)."""
    n = normalize(atomic)
    for kind, key, name, disp, why in FLAG_DISPOSITIONS:
        if kind == "prefix" and n.startswith(normalize(key)):
            return (name, disp, why)
        if kind == "contains" and normalize(key) in n:
            return (name, disp, why)
        if kind == "all" and all(normalize(k) in n for k in key):
            return (name, disp, why)
    return None


def split_flags_joinaware(cell):
    """Split a flag cell into distinct flags. The engine joins flags with '; '
    but some individual flags contain an internal '; '; those continuations
    start lowercase, so re-attach them to the flag they belong to. Lifted from
    generate_review_report.split_flags_joinaware."""
    parts = []
    for p in str(cell).split(";"):
        p = p.strip()
        if not p:
            continue
        if parts and p[:1].islower():
            parts[-1] = parts[-1] + "; " + p
        else:
            parts.append(p)
    return parts


def atomic_flags(*cells):
    """All atomic flags across the given flag cells (skips blanks/NaN)."""
    out = []
    for cell in cells:
        if cell is None:
            continue
        text = str(cell).strip()
        if not text or text.lower() == "nan":
            continue
        out.extend(split_flags_joinaware(text))
    return out


def build_review_cell(flags, uncatalogued_counter=None, disp_counter=None):
    """Build the review-column cell text for one row's atomic flags."""
    lines = []
    for f in flags:
        hit = classify_flag(f)
        if hit is None:
            if uncatalogued_counter is not None:
                uncatalogued_counter[normalize(f)] += 1
            lines.append(f"⚠ UNCATALOGUED -> review :: {f}")
            continue
        name, disp, why = hit
        if disp_counter is not None:
            disp_counter[disp] += 1
        lines.append(f"{name} -> {disp} ({why})")
    # Join with ' || ' (single line) so the cell is readable without needing
    # Excel wrap-text turned on.
    return "  ||  ".join(lines)


# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------

def _load_via_copy(src):
    """Load the workbook from a temp copy so a live Excel/OneDrive lock does
    not block reading (shutil.copy2 succeeds even when direct open-for-write
    does not)."""
    fd, tmp = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    shutil.copy2(src, tmp)
    return load_workbook(tmp), tmp


def find_headers(ws):
    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    headers = {c.value: c.column for c in header_cells if c.value is not None}
    for needed in (ETHNIC_FLAG_COL, GENDER_FLAG_COL, SEXUAL_FLAG_COL):
        if needed not in headers:
            raise SystemExit(f"ERROR: column '{needed}' not found in sheet '{ws.title}'. "
                             f"Found: {list(headers)}")
    return headers


def drift_report(src_data_path):
    """Live-recompute the three flags per row and compare to the stored cells.
    Reports how many rows drift (stored != live). Best-effort; skips cleanly if
    the engine/taxonomy cannot be loaded."""
    try:
        import pandas as pd
        sys.path.insert(0, str(bootstrap.PROJECT_ROOT / "Engine_1_and_2" / "Pipeline"))
        import ethnic_taggerv3 as et
        from classify_pipeline import classify_row as pipeline_classify_row
        import Gender_SexID as gs

        import paths
        tax_path = paths.taxonomy_file()
        tax_df = pd.read_excel(tax_path, sheet_name=et.TAXONOMY_SHEET, dtype=str)
        taxonomy_entries = et.build_taxonomy(tax_df)
        # Read via a temp copy so a live Excel/OneDrive lock on the workbook
        # does not block the drift check (matches _load_via_copy).
        fd, dtmp = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        shutil.copy2(src_data_path, dtmp)
        try:
            data_df = pd.read_excel(dtmp, sheet_name=et.DATA_SHEET, dtype=str)
        finally:
            try:
                os.remove(dtmp)
            except OSError:
                pass
    except Exception as e:  # pragma: no cover - environment dependent
        print(f"  drift check skipped ({type(e).__name__}: {e})")
        return

    def norm_cell(v):
        # Treat blank/NaN cells as empty so a stored blank ("nan") does not
        # register as drift against the live engine's empty-string "".
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        s = str(v).strip()
        return normalize(s) if s.lower() != "nan" else ""

    drift = 0
    for _, row in data_df.iterrows():
        _, _, _, eflag = pipeline_classify_row(row, taxonomy_entries)
        _, gflag = gs.classify_gender(row)
        _, sflag = gs.classify_sexual(row)
        stored = (norm_cell(row.get(ETHNIC_FLAG_COL, "")),
                  norm_cell(row.get(GENDER_FLAG_COL, "")),
                  norm_cell(row.get(SEXUAL_FLAG_COL, "")))
        live = (norm_cell(eflag), norm_cell(gflag), norm_cell(sflag))
        if stored != live:
            drift += 1
    print(f"  rows where a stored flag differs from the live engine: {drift} / {len(data_df)}")
    if drift:
        print("  -> AM/AO/AP may be stale; consider re-running the classifier before the real write.")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--dry-run", action="store_true",
                    help="compute + report only; do not write any workbook")
    ap.add_argument("--out", default=None,
                    help="write to THIS path (a test copy) instead of the real "
                         "FR testing.xlsx. Use for the copy-first safety gate; the "
                         "real workbook is left untouched and no backup is made.")
    args = ap.parse_args()

    ds = bootstrap.dataset()
    print(f"[dataset] active: {ds.name}  ->  reads {ds.raw_file.name}, writes {ds.output_file.name}")

    src = ds.output_file
    if not src.exists():
        raise SystemExit(f"ERROR: workbook not found at {src}")

    wb, tmp = _load_via_copy(src)
    try:
        if ds.data_sheet not in wb.sheetnames:
            raise SystemExit(f"ERROR: sheet '{ds.data_sheet}' not found. Sheets: {wb.sheetnames}")
        ws = wb[ds.data_sheet]
        headers = find_headers(ws)
        eth_c, gen_c, sex_c = (headers[ETHNIC_FLAG_COL], headers[GENDER_FLAG_COL],
                               headers[SEXUAL_FLAG_COL])
        # APPEND at the first free column past the used range (mirrors
        # ethnic_taggerv3.main's output-column append). NEVER insert_cols --
        # inserting mid-sheet corrupts the AM:AQ Excel Table (Table177).
        append_at = ws.max_column + 1

        disp_counter = Counter()
        uncatalogued = Counter()
        review_cells = {}   # row index -> cell text
        flagged_rows = 0

        for r in range(2, ws.max_row + 1):
            flags = atomic_flags(ws.cell(r, eth_c).value,
                                 ws.cell(r, gen_c).value,
                                 ws.cell(r, sex_c).value)
            if not flags:
                continue
            flagged_rows += 1
            review_cells[r] = build_review_cell(flags, uncatalogued, disp_counter)

        total_instances = sum(disp_counter.values()) + sum(uncatalogued.values())
        print(f"Sheet '{ds.data_sheet}': {ws.max_row - 1} data rows, {flagged_rows} with >=1 flag.")
        print(f"Flag columns: {ETHNIC_FLAG_COL}={get_column_letter(eth_c)}, "
              f"{GENDER_FLAG_COL}={get_column_letter(gen_c)}, "
              f"{SEXUAL_FLAG_COL}={get_column_letter(sex_c)}")
        print(f"New review column APPENDED at {get_column_letter(append_at)} "
              f"(past the used range; Table177 untouched)")
        print(f"\nDisposition tally ({total_instances} flag instances):")
        for disp, n in disp_counter.most_common():
            print(f"  {disp:5}  {n}")
        print(f"\nCatalog coverage: {len(uncatalogued)} distinct UNCATALOGUED flag(s).")
        for flag, n in uncatalogued.most_common():
            print(f"  [{n}] {flag[:100]}")

        if args.dry_run:
            print("\n--- drift check (live engine vs stored cells) ---")
            drift_report(src)
            print("\nDRY RUN: nothing written.")
            return

        if uncatalogued:
            raise SystemExit("\nABORT: uncatalogued flags present. Extend FLAG_DISPOSITIONS "
                             "until dry-run coverage is clean before writing.")

        # Populate the appended column (header + per-row cells). No insert, so
        # every other cell/table/validation stays in place.
        ws.cell(row=1, column=append_at, value=REVIEW_COL_HEADER)
        for r, text in review_cells.items():
            ws.cell(row=r, column=append_at, value=text)

        if args.out:
            # Copy-first safety gate: write to a test file, leave the real
            # workbook untouched, no backup.
            out_path = Path(args.out)
            wb.save(out_path)
            print(f"\nTEST COPY written (real workbook untouched): {out_path}")
            print("Open this copy in Excel and confirm there is NO recovery prompt "
                  "before applying to the real file.")
            return

        # Real write: pre-check writability BEFORE the backup so a locked
        # workbook fails cleanly without orphaning a backup.
        try:
            with open(src, "r+b"):
                pass
        except PermissionError:
            raise SystemExit(
                f"\nERROR: {src.name} is open/locked (close it in Excel and re-run). "
                "Nothing was written; no backup created.")

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = src.with_name(f"{src.stem}.backup_{ts}{src.suffix}")
        shutil.copy2(src, backup)
        print(f"\nBackup written: {backup.name}")
        try:
            wb.save(src)
        except PermissionError:
            raise SystemExit(
                f"\nERROR: could not write {src.name} -- it is open/locked. Close it "
                f"and re-run. Your data is untouched; backup exists at {backup.name}.")
        print(f"Wrote column '{REVIEW_COL_HEADER}' at {get_column_letter(append_at)} "
              f"for {len(review_cells)} flagged rows.")
        print(f"Saved: {src}")
    finally:
        try:
            os.remove(tmp)
        except OSError:
            pass


if __name__ == "__main__":
    main()
