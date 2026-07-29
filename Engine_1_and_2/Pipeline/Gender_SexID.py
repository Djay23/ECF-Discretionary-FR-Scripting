import re
import sys
import time
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))  # Engine 1 and 2
import bootstrap

from ethnic_taggerv3 import (
    DATA_SHEET,
    normalize_text,
    get_body_and_name_texts,
    matches_any,
    is_negated,
    is_example_mention,
    _echoes_org_name,
    parse_raw_org_name,
    body_names_a_population,
    _body_without_org_name,
)
from constants import ASPIRATIONAL_PHRASES, IDENTITY_EXPANSION_DISCLAIMER_PATTERNS

from gender_constants import (
    # Gender identity labels
    GENDER_WOMEN_GIRLS, GENDER_MEN_BOYS, GENDER_TWO_SPIRIT,
    GENDER_OTHER, GENDER_MULTIPLE, GENDER_GENERAL_POP,
    # Gender output columns
    OUTPUT_GENDER, OUTPUT_GENDER_FLAG,
    # Gender term data
    IDENTITY_KEY_TO_LABEL, IDENTITY_KEY_SHORT_LABEL,
    GENDER_TERM_PATTERNS,
    BARE_QUEER_PATTERN, BARE_TRANS_PATTERN, GENDER_QUEER_CONTEXT,
    UMBRELLA_ACRONYM_PATTERN,
    # Gender flags
    FLAG_ASPIRATIONAL, FLAG_NEGATION,
    FLAG_ORG_NAME,
    # Org-name data
    ORG_NAME_CONTEXT_PATTERNS,
    # Generalizable silent-body name rule
    GENDER_SILENT_NAME_WOMEN_PATTERN, GENDER_SILENT_NAME_MEN_PATTERN,
    SEXUAL_SILENT_NAME_PATTERN,
    # Incidental family-context guard
    RELATIONAL_MALE_GUARD_PATTERNS,
    RELATIONAL_MALE_SERVED_AFTER_PATTERNS, RELATIONAL_MALE_SERVED_BEFORE_PATTERNS,
    FAMILY_LEFT_BEHIND_BEFORE_PATTERNS, FAMILY_LEFT_BEHIND_AFTER_PATTERNS,
    # Gender-neutral org names (Boys & Girls Club / Big Brothers Big Sisters)
    GENDER_NEUTRAL_ORG_PATTERNS,
    # Curated known gender-serving org names (last-resort lookup)
    GENDER_ORG_NAME_MAP,
    # Sexual identity labels
    SEXUAL_2SLGBTQIA, SEXUAL_GENERAL_POP,
    # Sexual output columns
    OUTPUT_SEXUAL, OUTPUT_SEXUAL_FLAG,
    # Sexual term data
    SEXUAL_GENDER_DIVERSE_PATTERNS, SEXUAL_GENDER_DIVERSE_KEY_ONLY_PATTERNS,
    SEXUAL_ORIENTATION_PATTERNS,
    # Sexual flags
    SFLAG_NEGATION, SFLAG_GENDER_TERM, SFLAG_ASPIRATIONAL,
)

"""
Gender_SexID.py
---------------
Gender Identity + Sexual Identity classifiers — standalone script.

Reads:   "Data Sheets/FR testing.xlsx"  →  sheet "Discretionary Funding Requests"
Writes:  four columns back into the same workbook (openpyxl header-lookup):
           Gender Id - FR9       Gender Classification Flag
           Sexual Id - FR10      Sexual Classification Flag

Architecture:
    All data (labels, patterns, flag strings) lives in gender_constants.py.
    Generic text helpers (normalize_text, is_negated, etc.) are imported
    from ethnic_taggerv3.py — not copied.

Public API (imported by generate_review_report.py):
    classify_gender(row)  → (label: str, flag: str)
    classify_sexual(row)  → (label: str, flag: str)
    GENDER_GENERAL_POP, GENDER_MULTIPLE
    SEXUAL_2SLGBTQIA, SEXUAL_GENERAL_POP

Pipeline — Gender:
    1. get_body_and_name_texts() → normalize → extract_gender_candidates() on
       the body; a name-only signal degrades to General + FLAG_ORG_NAME
    2. resolve_gender() : 0 keys → General Pop, 1 → that group, 2+ → Multiple

Pipeline — Sexual:
    1. same body/name split → extract_sexual_candidates()
    2. resolve_sexual() : any guarded body signal → 2SLGBTQIA+, else General Pop
       (name-only signal → General Pop + FLAG_ORG_NAME)

To Run:
    python Gender_SexID.py
"""

# ===========================================================================
# ORG-NAME CONTEXT DETECTOR
# ===========================================================================

def is_org_name_context(text):
    """
    Return True when any ORG_NAME_CONTEXT_PATTERNS fires on normalized text,
    indicating a gender/sexual term appears inside an organizational construct.
    Classification is kept; caller appends FLAG_ORG_NAME as an annotation.
    """
    return matches_any(ORG_NAME_CONTEXT_PATTERNS, text)

SILENT_NAME_FLAG_TEXT = "Classified from organization name (silent body) - verify served population"

# ===========================================================================
# INCIDENTAL FAMILY-CONTEXT GUARD
# ===========================================================================

def _is_family_left_behind_mention(text, start, end, window=60):
    """Weak 'family_context' occurrence -- mirrors ethnic_taggerv3.infer_role's
    weak-role frame checks (org_name/provider/example/aspirational). True
    when a relational-male kinship term (fathers/dads/brothers/sons) sits in
    a frame describing an ABSENT/left-behind family member (e.g. "...their
    fathers... remain in Ukraine") rather than naming the served population
    directly (e.g. "a fathers group for new parents", which matches
    neither frame and stays served)."""
    before = text[max(0, start - window):start]
    after = text[end:end + window]
    return (
        matches_any(FAMILY_LEFT_BEHIND_BEFORE_PATTERNS, before)
        or matches_any(FAMILY_LEFT_BEHIND_AFTER_PATTERNS, after)
    )

def gender_org_lookup(*texts):
    """Last-resort curated known-org lookup (see GENDER_ORG_NAME_MAP). Returns
    the mapped gender label if any curated org name appears in the given
    text(s), else None. Consulted only when the body carries no served gender
    signal, mirroring the ethnic ORG_NAME_ETHNICITY_MAP fallback."""
    blob = " ".join(t for t in texts if t)
    for org_name, label in GENDER_ORG_NAME_MAP.items():
        if re.search(r"\b" + re.escape(org_name) + r"\b", blob, re.IGNORECASE):
            return label
    return None

def _relational_male_in_served_frame(text, start, end, window=40):
    """True when a relational-male noun (father/dad/brother/son) at
    text[start:end] is framed as the SERVED population -- immediately followed
    by a service noun ("a fathers group") or preceded by a serve verb ("for
    fathers", "supporting dads"). Only such framed occurrences count as served;
    a bare/incidental relational mention ("their fathers were not allowed to
    leave", "dad time", "her little brother kenny") does not."""
    before = text[max(0, start - window):start]
    after = text[end:end + window]
    return (
        matches_any(RELATIONAL_MALE_SERVED_BEFORE_PATTERNS, before)
        or matches_any(RELATIONAL_MALE_SERVED_AFTER_PATTERNS, after)
    )

def classify_gender_from_raw_name(raw_name):
    """
    Generalizable silent-body name rule — gender
    axis. Only consulted when the body carries NO gender-term mention at
    all (not even a weak org-name echo — see classify_gender). Classifies
    from women/men terms in the RAW funding-request account name; both
    present (e.g. "Boys & Girls Club") -> General. Returns (label, flag) or
    None if the name carries no gender term either.
    """
    org = parse_raw_org_name(raw_name)
    if not org:
        return None
    org_norm = normalize_text(org)
    has_women = bool(re.search(GENDER_SILENT_NAME_WOMEN_PATTERN, org_norm, re.IGNORECASE))
    has_men = bool(re.search(GENDER_SILENT_NAME_MEN_PATTERN, org_norm, re.IGNORECASE))
    if has_women and has_men:
        return GENDER_GENERAL_POP, SILENT_NAME_FLAG_TEXT
    if has_women:
        return GENDER_WOMEN_GIRLS, SILENT_NAME_FLAG_TEXT
    if has_men:
        return GENDER_MEN_BOYS, SILENT_NAME_FLAG_TEXT
    return None

def classify_sexual_from_raw_name(raw_name):
    """Generalizable silent-body name rule — sexual-identity axis. Mirrors
    classify_gender_from_raw_name; returns (label, flag) or None."""
    org = parse_raw_org_name(raw_name)
    if not org:
        return None
    org_norm = normalize_text(org)
    if re.search(SEXUAL_SILENT_NAME_PATTERN, org_norm, re.IGNORECASE):
        return SEXUAL_2SLGBTQIA, SILENT_NAME_FLAG_TEXT
    return None

# ===========================================================================
# GENDER IDENTITY — extraction + resolver
# ===========================================================================

def extract_gender_candidates(text, name_text=""):
    """
    Scan normalized text for gender-identity terms.

    Returns
    -------
    keys: set[str]  — accepted identity keys
    org_echo_keys: set[str]  — subset of keys whose match merely echoes the
        org's own name inside the body (e.g. an org restating its own name
        in its description) rather than describing a served population.
    family_context_keys: set[str]  — subset of keys whose match is only a
        relational male noun (father/dad/brother/son) with no corroborating
        direct male term (man/men/boy/male) in the body — an incidental
        family mention (e.g. "...their fathers... remain in Ukraine", or a
        volunteer who is a father) rather than the served population.
        Disjoint from org_echo_keys.
    flags_out: set[str]  — annotation flag strings triggered by this text
    any_negated : bool — True if any negation encountered (→ FLAG_NEGATION)
    """
    keys = set()
    keys_with_served_occurrence = set()
    keys_with_org_echo_occurrence = set()
    flags_out = set()
    any_negated = False

    # Guard 2 — gender-neutral org names. Spans of "Boys & Girls Club" /
    # "Big Brothers Big Sisters" in the body: gender words inside these name
    # spans identify the organization, not a served population, and are
    # skipped below (a genuine served mention elsewhere in the body is
    # unaffected).
    neutral_org_spans = [
        (m.start(), m.end())
        for pat in GENDER_NEUTRAL_ORG_PATTERNS
        for m in re.finditer(pat, text, re.IGNORECASE)
    ]

    # Standard (unambiguous) patterns.
    # Scan ALL occurrences (not just the first) so that when a term appears
    # more than once -- e.g. an org restates its own name ("<Org> Shelter
    # provides...") AND separately describes who it serves ("...for women
    # fleeing violence") -- a later served occurrence rescues the key even
    # if an earlier occurrence was an org-name echo (or an incidental
    # family-context mention).
    for pattern, identity_key, extra_flag_key in GENDER_TERM_PATTERNS:
        for m in re.finditer(pattern, text, re.IGNORECASE):
            term = m.group(0)
            # Guard 2: a gender word inside a gender-neutral org name
            # ("Boys & Girls Club", "Big Brothers Big Sisters") names the
            # organization, not who is served -- skip this occurrence.
            if any(s <= m.start() and m.end() <= e for s, e in neutral_org_spans):
                continue
            if is_negated(term, text):
                any_negated = True
                continue
            if is_example_mention(term, text):
                continue
            keys.add(identity_key)
            is_org_echo = bool(name_text) and _echoes_org_name(text, m.start(), m.end(), name_text)
            # Guard 1: a bare relational male noun (father/dad/brother/son) is
            # not a served signal on its own. It counts as served ONLY if a
            # direct male term (man/men/boy/male/...) also occurs in the body
            # -- that direct occurrence lands the key in
            # keys_with_served_occurrence via the else branch. A relational
            # occurrence with no such corroboration adds the key but no served
            # evidence, so it falls into the family-context weak bucket below.
            # (This subsumes the older left-behind-frame-only demotion; the
            # narrow FAMILY_LEFT_BEHIND check is no longer needed here.)
            is_relational_male = pattern in RELATIONAL_MALE_GUARD_PATTERNS
            if is_org_echo:
                keys_with_org_echo_occurrence.add(identity_key)
            elif is_relational_male and not _relational_male_in_served_frame(text, m.start(), m.end()):
                pass  # weak: relational male outside a served frame
            else:
                keys_with_served_occurrence.add(identity_key)

    weak_keys = keys - keys_with_served_occurrence
    org_echo_keys = weak_keys & keys_with_org_echo_occurrence
    # Remaining weak keys (relational-male-only, not an org echo) → family
    # context: present but not confirmed as the served population.
    family_context_keys = weak_keys - org_echo_keys

    # Bare "queer" — ambiguity check
    # \bqueer\b fires on "queer youth" AND on "gender queer community".
    # Suppress the ambiguity flag when "queer" is part of "gender queer".
    for m in re.finditer(BARE_QUEER_PATTERN, text, re.IGNORECASE):
        term = m.group(0)
        if is_negated(term, text) or is_example_mention(term, text):
            any_negated = True
            continue
        prefix = text[max(0, m.start() - 8) : m.start()].rstrip()
        if prefix.lower().endswith("gender") or prefix.lower().endswith("gender-"):
            continue  # already captured by the "gender queer" standard pattern
        keys.add("genderqueer")

    # Bare "trans" — false-positive risk
    # \btrans\b cannot match inside "transgender" (no \b between 's' and 'g').
    for m in re.finditer(BARE_TRANS_PATTERN, text, re.IGNORECASE):
        term = m.group(0)
        if is_negated(term, text) or is_example_mention(term, text):
            any_negated = True
            continue
        keys.add("transgender")

    # 2SLGBTQIA+ umbrella acronym — implies gender-diverse identity on the gender axis.
    # "2S" prefix additionally adds two_spirit (Two-Spirit is explicitly named in the acronym).
    # resolve_gender short-circuits lgbtq_umbrella → GENDER_MULTIPLE regardless of key count.
    for m in re.finditer(UMBRELLA_ACRONYM_PATTERN, text, re.IGNORECASE):
        term = m.group(0)
        if is_negated(term, text) or is_example_mention(term, text):
            any_negated = True
            continue
        keys.add("lgbtq_umbrella")
        if term.lower().startswith("2s"):
            keys.add("two_spirit")
        # FLAG_UMBRELLA_ACRONYM removed entirely;
        # the "Multiple: 2SLGBTQIA+ umbrella..." text in resolve_gender
        # already tells the reviewer the source, so this was pure duplicate
        # noise. Classification (the "lgbtq_umbrella" key) is unaffected.

    # FLAG_AMBIGUOUS_TERM removed entirely.
    # femme/masc/butch/stud are gender-coded slang; femme->women_girls is
    # still a real classifying signal via GENDER_TERM_PATTERNS above
    # (French "femme" = woman) and is unaffected. AMBIGUOUS_CODED_TERMS
    # itself is left in gender_constants.py -- audit_evidence.py still
    # consults it independently for evidence-column display.

    return keys, org_echo_keys, family_context_keys, flags_out, any_negated

def resolve_gender(keys, flags_set, any_negated, aspirational):
    """
    0 keys          → General Population (No specific gender served)
    lgbtq_umbrella  → Multiple gender identities (always — acronym spans identities)
    1 other key     → that key's output group label
    2+ other keys   → Multiple gender identities  (flag lists which identities)
    Annotation flags are appended; they never change the branch.

    Aspirational is a contextual flag (verify-manually), not
    classification-relevant like negation/BIPOC/two-spirit — it only fires
    once the row has actually resolved to a specific group, not General.
    """
    flag_parts = sorted(flags_set)

    if any_negated:
        flag_parts.append(FLAG_NEGATION)

    if not keys:
        return GENDER_GENERAL_POP, "; ".join(flag_parts)

    if aspirational:
        flag_parts.append(FLAG_ASPIRATIONAL)

    # LGBTQ family acronym → always Multiple regardless of other keys present.
    # The acronym inherently spans multiple gender identities; name the source in the flag.
    if "lgbtq_umbrella" in keys:
        other_short = sorted(IDENTITY_KEY_SHORT_LABEL[k] for k in keys if k != "lgbtq_umbrella")
        if other_short:
            label_list = ["2SLGBTQIA+ umbrella"] + other_short
        else:
            label_list = ["2SLGBTQIA+ umbrella acronym"]
        flag_parts.insert(0, f"Multiple: {', '.join(label_list)}")
        return GENDER_MULTIPLE, "; ".join(flag_parts)

    # 'gender-diverse' is a distinct concept, not a single
    # identity — it always routes to Multiple, same short-circuit shape as
    # the 2SLGBTQIA+ umbrella acronym above.
    if "gender_diverse" in keys:
        other_short = sorted(
            IDENTITY_KEY_SHORT_LABEL[k] for k in keys if k != "gender_diverse"
        )
        label_list = ["gender-diverse"] + other_short
        flag_parts.insert(0, f"Multiple: {', '.join(label_list)}")
        return GENDER_MULTIPLE, "; ".join(flag_parts)

    if len(keys) == 1:
        return IDENTITY_KEY_TO_LABEL[next(iter(keys))], "; ".join(flag_parts)

    # 2+ concrete, non-diverse keys (e.g.
    # men/boys + women/girls, non-binary + transgender) → Multiple with NO
    # primary flag; the label itself already says "Multiple", and which
    # identities combined isn't ambiguous enough to need a review flag.
    # Unlike the lgbtq_umbrella and gender_diverse branches above (kept),
    # every key reaching here is an unambiguous named identity.
    return GENDER_MULTIPLE, "; ".join(flag_parts)


def classify_gender(row):
    """Public entry point — gender identity classification for a single row.

    A signal must be corroborated in the body to classify; a name-only signal
    degrades to General Population + FLAG_ORG_NAME. A body mention that only
    echoes the org's own name (an org restating its own name in its
    description) is not corroboration either — it degrades to General + a
    transparency note. Likewise, a bare relational-male noun that only
    mentions an absent/left-behind family member (e.g. "...their fathers...
    remain in Ukraine") is not corroboration — General + a transparency note.
    """
    body_text, name_text = get_body_and_name_texts(row)
    if not (body_text + name_text).strip():
        return GENDER_GENERAL_POP, ""
    body = normalize_text(body_text)
    name = normalize_text(name_text)
    keys, org_echo_keys, family_context_keys, flags_set, any_negated = extract_gender_candidates(body, name_text=name)
    served_keys = keys - org_echo_keys - family_context_keys
    # Body is truly silent of served signal (no served keys, no
    # ambiguous-term/negation annotation) — only then does a name-only,
    # org-echo-only, or family-context-only signal matter.
    if not served_keys and not flags_set and not any_negated:
        # Last-resort curated known-org lookup: a specific org whose served gender is known from
        # outside the FR text, named in the body/name without a served signal.
        org_label = gender_org_lookup(body, name)
        if org_label is not None:
            return org_label, "Classified from a known organization name - verify served population"
        # Org-echo-only body: the org's own gender-named identity is restated
        # in an otherwise administrative body (e.g. a gender-named org whose
        # body only covers a server replacement). Per the org-name ladder this
        # is NOT a served signal but the org name IS the best available
        # evidence, so classify from the raw org name -- UNLESS the identity is
        # disclaimed ("beyond its original focus ...") or the body serves a
        # real population on another axis. The cross-axis gate runs on the
        # org-name-STRIPPED body so the echo itself is not miscounted as a
        # served population. Family-context-only mentions ("their fathers ...
        # remain in Ukraine") are NOT an org identity and fall through to
        # General below.
        disclaimed = matches_any(IDENTITY_EXPANSION_DISCLAIMER_PATTERNS, body)
        if org_echo_keys and not disclaimed and not body_names_a_population(
                _body_without_org_name(body, row.get("Funding Request Name", ""))):
            name_rule = classify_gender_from_raw_name(row.get("Funding Request Name", ""))
            if name_rule is not None:
                return name_rule
        notes = []
        if org_echo_keys:
            groups = sorted(IDENTITY_KEY_SHORT_LABEL[k] for k in org_echo_keys)
            notes.append(
                "Note (low priority): " + ", ".join(groups)
                + " mentioned via org-name self-reference in body - not classified; verify"
            )
        if family_context_keys:
            groups = sorted(IDENTITY_KEY_SHORT_LABEL[k] for k in family_context_keys)
            notes.append(
                "Note (low priority): " + ", ".join(groups)
                + " mentioned only in incidental family context (e.g. a relative left "
                "behind, or a relational term like father/brother outside a served frame) "
                "- not the served population; verify"
            )
        if notes:
            return GENDER_GENERAL_POP, "; ".join(notes)
        # Truly silent body (org_echo_keys and family_context_keys both
        # empty implies keys itself is empty here, since served_keys is
        # already empty in this branch) -- generalizable silent-body name
        # rule. Cross-axis gate: only apply it when the body is truly
        # administrative -- no served population on ANY axis (e.g. a body that
        # serves a named ethnic group but no gender: a real, gender-neutral
        # ethnic signal -- gender must stay General, not borrow "Women" from
        # the org name).
        if not body_names_a_population(body):
            name_rule = classify_gender_from_raw_name(row.get("Funding Request Name", ""))
            if name_rule is not None:
                return name_rule
        name_keys, _, _, _, _ = extract_gender_candidates(name)
        if name_keys:
            return GENDER_GENERAL_POP, FLAG_ORG_NAME
        return GENDER_GENERAL_POP, ""
    if served_keys and is_org_name_context(body):
        flags_set.add(FLAG_ORG_NAME)
    aspirational = matches_any(ASPIRATIONAL_PHRASES, body)
    return resolve_gender(served_keys, flags_set, any_negated, aspirational)

# ===========================================================================
# SEXUAL IDENTITY — extraction + resolver
# ===========================================================================

def extract_sexual_candidates(text):
    """
    Presence-only scan — any single guarded signal is enough.

    Returns
    -------
    found: bool — at least one signal survived guards
    found_gender_diverse: bool — at least one gender-diverse term survived guards
    found_gender_diverse_key: bool — the ambiguous "gender_diverse" umbrella
        term specifically (not trans/non-binary/two-spirit/etc.) survived
        guards (SFLAG_GENDER_TERM fires only
        when this is True, since the other gender-diverse-family terms
        unambiguously belong under 2SLGBTQIA+ already)
    any_negated : bool — at least one negation encountered
    example_suppressed : bool — at least one term was dropped ONLY because it
        appeared in an example/illustrative context ("...groups such as
        2SLGBTQIA+..."); the caller surfaces this as a
        transparency note instead of silently dropping it with no trace.
    """
    found_gender_diverse = False
    found_gender_diverse_key = False
    found_orientation = False
    any_negated = False
    example_suppressed = False

    for pattern in SEXUAL_GENDER_DIVERSE_PATTERNS:
        m = re.search(pattern, text, re.IGNORECASE)
        if not m:
            continue
        term = m.group(0)
        if is_negated(term, text):
            any_negated = True
            continue
        if is_example_mention(term, text):
            example_suppressed = True
            continue
        found_gender_diverse = True
        if pattern in SEXUAL_GENDER_DIVERSE_KEY_ONLY_PATTERNS:
            found_gender_diverse_key = True

    for pattern in SEXUAL_ORIENTATION_PATTERNS:
        m = re.search(pattern, text, re.IGNORECASE)
        if not m:
            continue
        term = m.group(0)
        if is_negated(term, text):
            any_negated = True
            continue
        if is_example_mention(term, text):
            example_suppressed = True
            continue
        found_orientation = True

    found = found_gender_diverse or found_orientation
    return found, found_gender_diverse, found_gender_diverse_key, any_negated, example_suppressed

def resolve_sexual(found, found_gender_diverse_key, any_negated, aspirational, example_suppressed=False):
    """
    found → 2SLGBTQIA+; else → General Population.
    SFLAG_GENDER_TERM fires only when the ambiguous "gender_diverse" umbrella
    term specifically contributed — the other
    gender-diverse-family terms (trans/non-binary/two-spirit/...) unambiguously
    belong under 2SLGBTQIA+ already, so flagging every one of those was noise.
    Flags are annotations only — they never change the branch.

    Aspirational is a contextual flag — it only fires once the row
    has actually resolved to 2SLGBTQIA+, not General.
    When the only signal was example-suppressed, keep General but
    name what was dropped instead of silently discarding it with no trace.
    """
    flag_parts = []

    if found:
        if found_gender_diverse_key:
            flag_parts.append(SFLAG_GENDER_TERM)
        if any_negated:
            flag_parts.append(SFLAG_NEGATION)
        if aspirational:
            flag_parts.append(SFLAG_ASPIRATIONAL)
        return SEXUAL_2SLGBTQIA, "; ".join(flag_parts)

    if example_suppressed:
        flag_parts.append(
            "Note (low priority): 2SLGBTQIA+ mentioned in example/illustrative "
            "context - not classified; verify"
        )
    if any_negated:
        flag_parts.append(SFLAG_NEGATION)
    return SEXUAL_GENERAL_POP, "; ".join(flag_parts)

def classify_sexual(row):
    """Public entry point — sexual identity classification for a single row.

    A signal must be corroborated in the body to classify; a name-only
    signal degrades to General Population + FLAG_ORG_NAME.
    """
    body_text, name_text = get_body_and_name_texts(row)
    if not (body_text + name_text).strip():
        return SEXUAL_GENERAL_POP, ""
    body = normalize_text(body_text)
    name = normalize_text(name_text)
    found, found_gender_diverse, found_gender_diverse_key, any_negated, example_suppressed = extract_sexual_candidates(body)
    # Body is truly silent (no signal, no negation, no example-suppressed
    # mention) — only then does a name-only signal matter.
    if not found and not any_negated and not example_suppressed:
        # Truly silent body -- generalizable silent-body name rule.
        # Cross-axis gate: only apply it when the body is truly
        # administrative -- no served population on ANY axis.
        if not body_names_a_population(body):
            name_rule = classify_sexual_from_raw_name(row.get("Funding Request Name", ""))
            if name_rule is not None:
                return name_rule
        name_found, *_ = extract_sexual_candidates(name)
        if name_found:
            return SEXUAL_GENERAL_POP, FLAG_ORG_NAME
        return SEXUAL_GENERAL_POP, ""
    aspirational = matches_any(ASPIRATIONAL_PHRASES, body)
    label, flag = resolve_sexual(found, found_gender_diverse_key, any_negated, aspirational, example_suppressed)
    if found and is_org_name_context(body):
        flag = "; ".join(p for p in [FLAG_ORG_NAME, flag] if p)
    return label, flag

# ===========================================================================
# main() — writes both classifiers to FR testing.xlsx in one workbook pass
# ===========================================================================

def main():
    start_time = time.time()

    ds = bootstrap.dataset()
    print(f"[dataset] active: {ds.name}  ->  reads {ds.raw_file.name}, writes {ds.output_file.name}")

    # Reads output_file, not raw_file: this stage appends to what
    # ethnic_taggerv3.main() (stage 1) already wrote, so it must run after it.
    funding_filepath = ds.output_file

    print(f"Loading funding requests from: {funding_filepath}")
    if not funding_filepath.exists():
        print(f"Error: data workbook not found at '{funding_filepath}'.")
        sys.exit(1)
    try:
        data_df = pd.read_excel(funding_filepath, sheet_name=DATA_SHEET, dtype=str)
    except Exception as e:
        print(f"Error loading data sheet '{DATA_SHEET}' from '{funding_filepath}': {e}")
        sys.exit(1)

    print(f"Data rows: {len(data_df)}")

    for col in [OUTPUT_GENDER, OUTPUT_GENDER_FLAG, OUTPUT_SEXUAL, OUTPUT_SEXUAL_FLAG]:
        if col not in data_df.columns:
            data_df[col] = ""

    stats = {
        "-Gender → Women & Girls": 0, "-Gender → Men & Boys": 0, "-Gender → Two Spirit": 0,
        "-Gender → Other": 0, "-Gender → Multiple": 0, "-Gender → General": 0,
        "-Sexual → 2SLGBTQIA": 0, "-Sexual → General": 0,
        "-Flagged → Gender": 0, "-Flagged → Sexual": 0,
    }

    for idx, row in data_df.iterrows():
        g_label, g_flag = classify_gender(row)
        s_label, s_flag = classify_sexual(row)

        data_df.at[idx, OUTPUT_GENDER]      = g_label
        data_df.at[idx, OUTPUT_GENDER_FLAG] = g_flag
        data_df.at[idx, OUTPUT_SEXUAL]      = s_label
        data_df.at[idx, OUTPUT_SEXUAL_FLAG] = s_flag

        if   g_label == GENDER_WOMEN_GIRLS: stats["-Gender → Women & Girls"] += 1
        elif g_label == GENDER_MEN_BOYS:    stats["-Gender → Men & Boys"]    += 1
        elif g_label == GENDER_TWO_SPIRIT:  stats["-Gender → Two Spirit"]  += 1
        elif g_label == GENDER_OTHER:       stats["-Gender → Other"]       += 1
        elif g_label == GENDER_MULTIPLE:    stats["-Gender → Multiple"]    += 1
        else:                               stats["-Gender → General"]     += 1

        if s_label == SEXUAL_2SLGBTQIA:     stats["-Sexual → 2SLGBTQIA"]  += 1
        else:                               stats["-Sexual → General"]     += 1

        if g_flag: stats["-Flagged → Gender"]  += 1
        if s_flag: stats["-Flagged → Sexual"]  += 1

    # Write-back via openpyxl header-lookup (never assumes column positions)
    wb = load_workbook(funding_filepath)
    ws = wb[DATA_SHEET]
    headers = {cell.value: cell.column for cell in ws[1]}

    for col_name in [OUTPUT_GENDER, OUTPUT_GENDER_FLAG, OUTPUT_SEXUAL, OUTPUT_SEXUAL_FLAG]:
        if col_name not in headers:
            new_col = ws.max_column + 1
            ws.cell(row=1, column=new_col, value=col_name)
            headers[col_name] = new_col

    for i, (idx, _row) in enumerate(data_df.iterrows(), start=2):
        ws.cell(row=i, column=headers[OUTPUT_GENDER],      value=data_df.at[idx, OUTPUT_GENDER])
        ws.cell(row=i, column=headers[OUTPUT_GENDER_FLAG], value=data_df.at[idx, OUTPUT_GENDER_FLAG])
        ws.cell(row=i, column=headers[OUTPUT_SEXUAL],      value=data_df.at[idx, OUTPUT_SEXUAL])
        ws.cell(row=i, column=headers[OUTPUT_SEXUAL_FLAG], value=data_df.at[idx, OUTPUT_SEXUAL_FLAG])

    wb.save(funding_filepath)

    print("\nResults:")
    for k, v in stats.items():
        print(f"  {k}: {v}")
    print(f"\nOutput written to: {funding_filepath}")
    elapsed = time.time() - start_time
    print(f"Gender + Sexual classification completed in {elapsed:.1f} seconds.")

if __name__ == "__main__":
    main()
