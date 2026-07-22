import sys
import re
import pandas as pd
import time
from collections import defaultdict

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path

"""
ethnic_tagger_v3.py
--------------------
Rule-based ethnic origin classifier for ECF Discretionary
Funding Requests.

Reads:
    "Taxonomy - Definitions.xlsx" -> sheet "Ethnic and Cultural Origins"
    "Discretionary FR working -2025 (Oreva).xlsx" -> sheet "Discretionary Funding Requests"

Aggregates matches across all 4 input columns (Final_Project_Description,
Final_Summary_Description, Purpose, Funding Request Name), then resolves
using the case hierarchy below.

To Run:
    python ethnic_taggerv3.py "C:\\Users\\oadode\\OneDrive - Edmonton Community Foundation\\Desktop\\Discretionary FR Scripting\\ECF-Discretionary-FR-Scripting\\Taxonomy - Definitions.xlsx" "C:\\Users\\oadode\\OneDrive - Edmonton Community Foundation\\Desktop\\Discretionary FR Scripting\\ECF-Discretionary-FR-Scripting\\FR testing.xlsx"

Case coverage (see README):
    1.  Exact match (deepest taxonomy term)
    2.  Level 2 match (subregion)
    3.  Level 1 match (broad category)
    4.  Structured phrase ("South African" etc.) -> pattern rules
    5.  Country/nationality IN taxonomy -> handled by case 1-3
    6.  Country/nationality NOT in taxonomy -> country map
    7.  "from <country>" phrasing -> country map (phrase variant)
    8.  Multiple groups -> Multiple Ethnic and Cultural Origins
    9.  BIPOC (context-aware) -> Multiple Ethnic and Cultural Origins
    9b. Broad identity labels (Black, Arab, Jewish, etc.) -> Other Ethnic
        and Cultural Origins
    10. Known organization name lookup (Bent Arrow, Treaty 6, etc.)
        -> only consulted if result would otherwise be General Population
    11. "Grassroots" -> ethnic-indicating ONLY if paired with another
        ethnic keyword; otherwise ignored entirely
    12. General Population (fallback / no signal)

Also handles:
    - Context override ("beyond its original focus on X")
        - Historical reference ("historically served X")
        - Expansion Phrases ("")
    - Negation ("not targeting X", "does not serve X")
    - Aspirational language ("plans to expand to X") -> flagged for review
    - Example mentions ("such as X", "including X") -> not classified
    - Strict word-boundary matching (no "blacksmith" -> "black")
"""

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))  # Engine 1 and 2
import bootstrap 

# Semantic-similarity fallback (see semantic_fallback.py).
try:
    import semantic_fallback
    SEMANTIC_AVAILABLE = True
except ImportError:
    SEMANTIC_AVAILABLE = False

from constants import (
    MULTIPLE_ETHNIC, OTHER_ETHNIC, GENERAL_POP,
    BIPOC_KEYWORDS, BIPOC_PROGRAM_NAME_AFTER_PATTERNS, TAXONOMY_KEYWORD_STOPLIST,
    AMBIGUOUS_EQUITY_WORDS, BROAD_IDENTITY_KEYWORDS,
    ALWAYS_MULTIPLE_COMPOUNDS, ORG_NAME_ETHNICITY_MAP,
    PATTERN_RULES, COUNTRY_REGION_MAP,
    EXPANSION_PHRASES, HISTORICAL_PHRASES, NEGATION_PHRASES,
    ETHNIC_ANNOTATION_NEGATION_PHRASES,
    ASPIRATIONAL_PHRASES, EXAMPLE_PHRASES, EXPERT_ROLE_PHRASES,
    SERVING_CONTEXT_WORDS, IDENTITY_PHRASE_REWRITES,
    DIRECTIONAL_AFRICAN_PREFIXES, CASE_13_PATTERNS,
    DEMONYM_SUFFIXES, NON_ETHNIC_LEADING_WORDS,
    ROLE_ORG_NAME_BEFORE_PATTERNS, ROLE_ORG_NAME_AFTER_PATTERNS,
    ROLE_PROVIDER_BEFORE_PATTERNS, ROLE_PROVIDER_AFTER_PATTERNS,
    SERVED_FRAME_BEFORE_PATTERNS, SERVED_FRAME_CONTINUATION_PATTERNS,
    ROLE_TOPIC_AFTER_PATTERNS, INDIGENOUS_CONTEXT_RESCUE_PATTERNS,
    ROLE_TOPIC_KEEP_AFTER_PATTERNS, ROLE_TOPIC_KEEP_BEFORE_PATTERNS,
    ASPIRATIONAL_WIDE_WINDOW,
    ROLE_SETTING_BEFORE_PATTERNS, ROLE_SETTING_AFTER_PATTERNS,
    ROLE_ALLYSHIP_BEFORE_PATTERNS,
)

# gender_constants.py is a pure-data module (no imports of its own) so
# pulling in its taxonomy-independent term patterns here creates no
# circular dependency. Used only by body_names_a_population() below.
from gender_constants import GENDER_TERM_PATTERNS, SEXUAL_ORIENTATION_PATTERNS

# =============
# CONFIGURATION 
# =============

from dataset_config import active_config
_name, _cfg = active_config()
TAXONOMY_SHEET = _cfg["taxonomy_sheet"]
DATA_SHEET = _cfg["data_sheet"]

TAXONOMY_ENTRY1 = "Ethnic and Cultural Origins Level 1"
TAXONOMY_ENTRY2 = "Ethnic and Cultural Origins Level 2"
TAXONOMY_ENTRY3 = "Ethnic and Cultural Origins Level 3"
TAXONOMY_ALL_TERMS = "All Terms"
TAXONOMY_SCOPE_NOTES = "Definitions and Scope Notes"

# Search priority order. All 4 are concatenated, but this order
# determines which match "wins" on ties (first column listed wins)
INPUT_COLS_PRIORITY = [
    "Final_Project_Description",
    "Final_Summary_Description",
    "Purpose",
    "Funding Request Name",
]

# Name/body split (Plan.md Phase 1, D1): a signal must be corroborated in the
# served-population body text to classify; a name-only signal degrades to
# General + flag. See get_body_and_name_texts().
BODY_COLS = ["Final_Project_Description", "Final_Summary_Description", "Purpose"]
NAME_COL  = "Funding Request Name"

OUTPUT_ETHNIC1 = "Ethnic 1 - FR6"
OUTPUT_ETHNIC2 = "Ethnic 2 - FR7"
OUTPUT_ETHNIC3 = "Ethnic 3 - FR8"
OUTPUT_FLAG = "Classification Flag"
OUTPUT_SEMANTIC = "Semantic Suggestion (REVIEW)"


# =======
# HELPERS
# =======

def clean_taxonomy_value(val):
    """Lowercase, strip, and remove the literal word 'origins'. Treats NaN/None as
    empty. Pandas reads blank Excel cells as float('nan'), so an explicit pd.isna() check is required here."""
    if val is None or (isinstance(val, float) and pd.isna(val)) or pd.isna(val):
        return ""
    val = str(val).strip()
    if val.lower() == "nan":
        return ""
    val = val.lower()
    val = val.replace("origins", "")
    return val.strip()

def safe_display(val):
    """Return the original display value for level1/level2/level3,
    treating NaN as empty string instead of the literal 'nan'."""
    if val is None or pd.isna(val):
        return ""
    val = str(val).strip()
    return "" if val.lower() == "nan" else val

def extract_match_keyword(raw_keyword):
    """Trim a taxonomy label down to the bare word(s) worth searching
    for in funding-request text: "black, not otherwise specified" ->
    "black", "indian (india)" -> "indian"."""
    keyword = raw_keyword.split(",")[0]
    keyword = re.sub(r"\s*\([^)]*\)\s*$", "", keyword)
    return keyword.strip()

def normalize_text(text):
    """Lowercase, strip punctuation, collapse whitespace. Used on the
    funding-request free text columns (not the taxonomy sheet, which
    uses clean_taxonomy_value instead)."""
    if pd.isna(text) or not isinstance(text, str):
        return ""
    text = text.lower()
    text = re.sub(r'[^\w\s]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

# UNCOMMENT : For implementation of alias check
# ------
# def apply_aliases(text):
#     for alias, canonical in KEYWORD_ALIASES.items():
#         text = re.sub(r'\b' + re.escape(alias) + r'\b', canonical, text)
#     return text

def matches_any(patterns, text):
    return any(re.search(p, text, re.IGNORECASE) for p in patterns)

def keyword_context_window(keyword, text, window=80):
    """Return the text immediately preceding the first occurrence of
    keyword, used to check for negation/example/override phrases that
    appear right before it."""
    pattern = re.compile(r'\b' + re.escape(keyword) + r'\b', re.IGNORECASE)
    m = pattern.search(text)
    if not m:
        return None
    start = max(0, m.start() - window)
    return text[start:m.start()]

def is_non_prefixed(keyword, text):
    """Return True if keyword appears immediately after 'non-' or 'non ' in text."""
    return bool(re.search(r'\bnon[\- ]' + re.escape(keyword), text, re.IGNORECASE))

def is_negated(keyword, text):
    snippet = keyword_context_window(keyword, text)
    return (
        (snippet is not None and matches_any(NEGATION_PHRASES, snippet))
        or is_non_prefixed(keyword, text)
    )

def is_example_mention(keyword, text):
    snippet = keyword_context_window(keyword, text)
    return snippet is not None and matches_any(EXAMPLE_PHRASES, snippet)

# ---------------------------------------------------------------------------
# Phase 2 — Evidence-role inference (Plan.md Fix 1 Phase 2)
#
# A matched identity term is "served" (strong, default) unless the text
# immediately around it frames it as an org name, a service provider, an
# example/aspirational mention, or a negation -- in which case it's "weak"
# and should not, on its own, drive a classification.
# ---------------------------------------------------------------------------

def _echoes_org_name(text, start, end, name_text, min_gram=3, slack=3):
    """True if THIS SPECIFIC match sits inside (or immediately touches) a
    verbatim restatement of the org's own name in the body -- e.g. the
    "women" in "Black Canadian Women in Action (BCW) is undertaking..." --
    a self-reference to the org's own name, not a served-population claim.

    Requires actual positional overlap with an occurrence of a contiguous
    min_gram-word run from name_text, NOT just proximity within some fixed
    character window -- normalize_text strips all punctuation (periods
    included), so a window-based check can't tell "the org's name and this
    term are in the same clause" from "the org happens to restate its own
    name in the very next, unrelated sentence" (e.g. HERizon Healing
    Society names itself at the start of almost every sentence; a
    "BIPOC girls and young women" mention two sentences earlier is not an
    echo just because "HERizon Healing Society" appears soon after it).
    """
    if not name_text:
        return False
    name_words = re.findall(r"[a-z0-9]+", name_text.lower())
    if len(name_words) < min_gram:
        return False
    for i in range(len(name_words) - min_gram + 1):
        gram_words = name_words[i:i + min_gram]
        if sum(len(w) for w in gram_words) < 8:
            continue
        gram_pattern = r"\s+".join(re.escape(w) for w in gram_words)
        for gm in re.finditer(gram_pattern, text, re.IGNORECASE):
            if gm.start() - slack <= end and start <= gm.end() + slack:
                return True
    return False

def _name_word_echo(span, name_text):
    """True if the matched span's bare word also appears, as a standalone
    word, in the funding-request NAME text (Plan.md Chunk G1 step 4 — echo
    demotion upgrade). This is a looser check than _echoes_org_name's
    contiguous multi-word positional overlap: it catches a single-word
    identity term (e.g. "Black") that the org's OWN NAME also carries (e.g.
    "Black Canadian Women in Action"), even when the body mention doesn't
    restate the full name verbatim near this occurrence (e.g. "...beyond
    its original focus on Black communities" — no "Canadian Women in
    Action" nearby, so _echoes_org_name alone would miss it).
    """
    if not name_text or not span:
        return False
    word = span.strip().lower()
    if not word:
        return False
    # Also try the singular form so "Somalis" (body) still echoes "Somali"
    # (name) and vice versa.
    candidates = {word}
    if word.endswith("s") and len(word) > 3:
        candidates.add(word[:-1])
    return any(
        re.search(r'\b' + re.escape(c) + r'\b', name_text, re.IGNORECASE)
        for c in candidates
    )

# Role returned when a term counts as SERVED only because the organization
# named ITSELF -- an org-name echo or a copula self-description. It behaves
# exactly like "served" for every classification decision (see is_served_role
# and extractors._candidate, which normalizes it back to "served" and records
# the provenance on the candidate as self_id=True), but it is weaker EVIDENCE:
# the text never actually states who is served, we inferred it from the org's
# identity. classify_row uses that provenance to keep such a row flagged, which
# is what the silent-body name rule used to guarantee before self-identification
# became a served signal.
SELF_ID_ROLE = "served_self_id"


def is_served_role(role):
    """True for both plain "served" and the self-identification variant.
    Callers deciding "does this occurrence corroborate a served population?"
    must use this rather than == "served", or a self-identified org silently
    stops corroborating."""
    return role in ("served", SELF_ID_ROLE)


# Copula immediately preceding the match: "<Org> ... is the only <Indigenous>
# Artist-Run Centre". Anchored to the END of the before-window so only a
# copula DIRECTLY in front of the term counts -- "IWDIMAA program from
# Cameroonian Association" has no copula and must not qualify.
_SELF_ID_COPULA = re.compile(
    r"\b(?:is|are|was|were)\s+(?:the\s+|a\s+|an\s+)?"
    r"(?:only\s+|first\s+|sole\s+|leading\s+|largest\s+)?$"
)

# Generic words in a funding-request name that identify no particular org.
_SELF_ID_NAME_STOPWORDS = {
    "funding", "request", "for", "the", "of", "and", "society", "association",
    "centre", "center", "foundation", "collective", "inc", "ltd", "edmonton",
    "alberta", "canada", "canadian", "community", "services", "service",
}


def _org_self_identifies(text, start, name_text, window=140):
    """True when the REQUESTING organization is describing ITSELF immediately
    before this match -- "Ociciwan Contemporary Art Collective is the only
    Indigenous Artist-Run Centre based in Edmonton".

    This is the counterpart to _name_word_echo for orgs whose name does NOT
    contain the identity word. An org named in an Indigenous language
    (Ociciwan, Niginan, amiskwaciy, Kihciy Askiy) can never match a name-string
    test, so its self-description would otherwise be demoted by the generic
    "<Ethnicity> Centre/Association" org-name pattern -- which exists to catch
    THIRD-PARTY orgs ("in partnership with the Council of India"), not the
    applicant describing itself. Without this, the name-string approach
    systematically misses exactly the orgs most likely to be Indigenous-serving.

    Requires BOTH a copula directly in front of the term AND a distinctive word
    from the requesting org's own name earlier in the same window, so a
    third-party org's self-description elsewhere in the text cannot trigger it.
    """
    if not name_text:
        return False
    before = text[max(0, start - window):start]
    if not _SELF_ID_COPULA.search(before):
        return False
    tokens = [t for t in re.findall(r"[a-z]+", name_text.lower())
              if len(t) > 3 and t not in _SELF_ID_NAME_STOPWORDS]
    return any(t in before for t in tokens)


def infer_role(text, start, end, name_text="", window=60):
    """Classify the evidence role of a match at text[start:end]:
    'served' (strong, default) or one of the weak roles
    'org_name' | 'provider' | 'example' | 'aspirational'.

    Negation is deliberately NOT a role here: negation stays annotation-only
    (a resolver-level architectural invariant — see resolver.py's module
    docstring and build_context_notes) rather than suppressing a candidate,
    unlike org-name/provider/example/aspirational framing which describe
    what the term refers to, not whether it's included or excluded.

    Historical/expansion framing ("beyond its original focus on X",
    "historically served X") is likewise NOT its own role (Plan.md Chunk G1
    step 3 — annotation-only invariant, see extract_context_signals). It
    demotes a candidate only indirectly: when it coincides with an org-name
    echo (step 2/4 below), the echo role wins; otherwise the match falls
    through to "served" like any other unclassified framing.
    """
    before = text[max(0, start - window):start]
    after  = text[end:end + window]
    span   = text[start:end]
    if matches_any(EXAMPLE_PHRASES, before):
        return "example"
    if matches_any(ASPIRATIONAL_PHRASES, before):
        return "aspirational"
    # Plan.md Chunk G6 — same aspirational-reach frames, wider window: a row
    # that restates one aspirational claim across two concatenated body
    # columns (e.g. "...hoping to double its reach in the Edmonton area. To
    # expand and hone programming to include more indigenous...") can put
    # the lead verb ("wants to"/"hoping to") more than `window` chars before
    # a LATER restatement of the same mention. Only consulted as a fallback
    # -- the narrow check above still wins when it applies.
    before_wide = text[max(0, start - ASPIRATIONAL_WIDE_WINDOW):start]
    if matches_any(ASPIRATIONAL_PHRASES, before_wide):
        return "aspirational"
    # Plan.md Chunk G7 — fictional/historical story-setting frame: the term
    # names the SETTING of a play/myth/themed event, not a real served
    # population (e.g. Theatre Prospero "based on the tale of an Egyptian
    # prince"; Arts On The Ave "A Byzantine Winter Festival"). Checked here,
    # ahead of org_name/provider, since a setting reference isn't an org
    # name or a provider role either -- it's simply not about a served
    # population at all. Reuses the weak "topic" role (same demotion
    # outcome as the G3 curriculum-topic case).
    if matches_any(ROLE_SETTING_BEFORE_PATTERNS, before) or matches_any(ROLE_SETTING_AFTER_PATTERNS, after):
        return "topic"
    # Self-identification beats the org-name patterns below (Ociciwan): those
    # patterns exist to demote THIRD-PARTY org names, but they also fire on the
    # applicant describing itself ("... is the only Indigenous Artist-Run
    # Centre"), where the identity IS the served population. Checked first so
    # the applicant's own claim wins; a genuine third party ("in partnership
    # with the Council of India") has no copula + own-name pair and still
    # demotes normally.
    if _org_self_identifies(text, start, name_text):
        return SELF_ID_ROLE
    if matches_any(ROLE_ORG_NAME_BEFORE_PATTERNS, before) or matches_any(ROLE_ORG_NAME_AFTER_PATTERNS, after):
        return "org_name"
    if matches_any(ROLE_PROVIDER_BEFORE_PATTERNS, before) or matches_any(ROLE_PROVIDER_AFTER_PATTERNS, after):
        return "provider"
    # Served-frame rescue (Plan.md Chunk G1 step 2) — MUST run before the
    # echo checks below. An explicit "for X"/"serving X" lead-in, or a
    # service-continuity phrase ("continues today"), means this occurrence
    # IS the served population even if it also happens to echo the org's
    # own name elsewhere in the same text (e.g. "Somali Canadian Cultural
    # Society ... serving the Somali community").
    if matches_any(SERVED_FRAME_BEFORE_PATTERNS, before) or matches_any(SERVED_FRAME_CONTINUATION_PATTERNS, text) \
            or matches_any(INDIGENOUS_CONTEXT_RESCUE_PATTERNS, text):
        return "served"
    # Allyship / reconciliation-ally frame (Plan.md Task 4 / G11): "an ally
    # to indigenous people", "allies of X", "in support of reconciliation"
    # names the org/speaker's relationship to the group, not the group as a
    # served population. Checked AFTER the served-frame rescue above, so a
    # genuinely served row (residential-school/ceremony context, or an
    # explicit "for X"/"serving X" lead-in) still wins and never reaches
    # this check. Group-agnostic, same as the setting/topic frames.
    if matches_any(ROLE_ALLYSHIP_BEFORE_PATTERNS, before):
        return "topic"
    # Org-name echo == SELF-IDENTIFICATION, not a non-claim.
    #
    # These two checks previously returned "org_name" (weak), on the theory
    # that an org restating its own name ("Black Canadian Women in Action
    # (BCW) is undertaking...") is a self-reference rather than a
    # served-population claim. The hand-audit in AUDITED_FR_GOLD.xlsx
    # contradicts that: an ethnicity-named org naming its own ethnicity IS
    # evidence of who it serves. Black Canadian Women in Action was demoted to
    # General on both counts -- the name echo AND the expansion phrasing
    # "beyond its original focus on Black communities" -- yet gold is Black.
    #
    # Measured over all 448 rows: treating these as served fixes that row and
    # regresses NOTHING. The demotion turned out to be near-inert, because
    # dedup() already rescues any group with one served mention elsewhere, so
    # it only ever changed the outcome when the echo was the ONLY mention --
    # exactly the self-identification case it was getting wrong.
    #
    # Third-party org names are still demoted, by the ROLE_ORG_NAME_*
    # patterns above; that path is untouched.
    if _echoes_org_name(text, start, end, name_text):
        return SELF_ID_ROLE
    if (matches_any(EXPANSION_PHRASES, before) or matches_any(HISTORICAL_PHRASES, before)) \
            and _name_word_echo(span, name_text):
        return SELF_ID_ROLE
    # Topic / consultation framing (Plan.md Chunk G3): a term immediately
    # followed by "perspectives" or "knowledge holders" describes a
    # CURRICULUM TOPIC being taught or a CONSULTED-PARTY mention, not the
    # served population -- e.g. "...teach eco-action strategies, and
    # Indigenous perspectives" (course content); "consult wildlife
    # biologists, conservation experts, and Indigenous knowledge holders"
    # (advisory role). The served-frame rescue above already runs first, so
    # any row where this framing coexists with an explicit served mention,
    # or a residential-school/Truth-and-Reconciliation/ceremony frame
    # elsewhere in the text, keeps "served" and never reaches this check.
    if matches_any(ROLE_TOPIC_AFTER_PATTERNS, after):
        return "topic"
    return "served"

def indigenous_topic_keep_role(text, start, end, window=100):
    """Plan.md Chunk G6 (hybrid policy) — ONLY consulted by extractors.py for
    a candidate already known to be North American Indigenous Origins (see
    extract_taxonomy_candidates/extract_pattern_candidates), and only when
    infer_role() above already returned "served" (the default). A term
    naming Indigenous knowledge/art/practice INTEGRATED as actual program
    content ("Indigenous wisdom", "Indigenous dance"), or an Indigenous
    community/nation named as an active partner ("in collaboration with
    Indigenous Nations"), is not a confident "served" claim but isn't
    clearly NOT one either -- resolver.py counts "topic_keep" as served for
    the outcome, and adds FLAG_INDIGENOUS_TOPIC_VERIFY only when no OTHER
    occurrence of the same group was a plain "served" mention.

    Deliberately NOT folded into infer_role() itself: "dance"/"art"/
    "knowledge" are far too common as the actual (non-Indigenous) org's own
    identity-adjacent words -- e.g. "Ukrainian Dance Festival" -- to use as
    a generic weak-role frame across every group; scoping to Indigenous-only
    candidates at the extractor call site (where level1 is known) avoids
    that cross-group collision entirely.
    """
    before = text[max(0, start - window):start]
    after  = text[end:end + window]
    if matches_any(ROLE_TOPIC_KEEP_BEFORE_PATTERNS, before) or matches_any(ROLE_TOPIC_KEEP_AFTER_PATTERNS, after):
        return "topic_keep"
    return None

def phrase_has_serving_context(pattern, text, window=100):
    for m in re.finditer(pattern, text, re.IGNORECASE):
        start = max(0, m.start() - window)
        end = min(len(text), m.end() + window)
        if matches_any(SERVING_CONTEXT_WORDS, text[start:end]):
            return True
    return False

# ============================================================
# TAXONOMY BUILDER (Case 1-3 source data)
# Sort: deepest first, then LONGEST keyword first within same depth. Prevents "African"
# from matching before "Southern and East African".
# ============================================================

# ============================================================
# TAXONOMY BUILDER (Case 1-3 source data)
#
# Pulled from Column D "All Terms", a flat concatenation of
# Level 1 + Level 2 + Level 3 with the literal word "Origins" acting
# as the delimiter after L1 and L2 (no separator after L3,
# since specific terms like "Somali" don't carry an "Origins" suffix).
#
# Example: "African OriginsSouthern and East African OriginsSomali"
#   -> split on "Origins" -> ["African", "Southern and East African", "Somali"]
#
# Falls back to reading Level 1/2/3 columns directly if All Terms is
# blank or missing for a given row, for resilience against a broken
# formula cell.
#
# Sort: deepest first, then LONGEST keyword first within the same
# depth. This is the fix from the production skeleton — prevents
# "African" from matching before "Southern and East African".
# ============================================================

def parse_all_terms(all_terms_value):
    """
    Split the 'All Terms' cell on the literal word 'Origins' into
    its component level strings, stripped of whitespace. Trailing chars removed.
    """
    if pd.isna(all_terms_value) or not isinstance(all_terms_value, str):
        return []
    parts = all_terms_value.split("Origins")
    parts = [p.strip() for p in parts if p.strip()]
    return parts

def build_taxonomy(tax_df):
    entries = []
    for _, row in tax_df.iterrows():
        all_terms_raw = row.get(TAXONOMY_ALL_TERMS, "")
        parts = parse_all_terms(all_terms_raw)

        if parts:
            # Primary path: parsed from All Terms column.
            # The split strips the literal word "Origins" from every
            # fragment, but in the real taxonomy, Level 1 and Level 2
            # display values usually INCLUDE the "Origins" suffix
            # (e.g. "African Origins", "Southern and East African Origins")
            # -- except entries like "Black, not otherwise specified" and
            # "Jewish" under "Other Ethnic and Cultural Origins", which
            # only have "Origins" on Level 1. Count actual occurrences of
            # "Origins" in the source string so we know how many leading
            # parts really need it re-appended, instead of assuming L1+L2
            # always do.
            origins_count = all_terms_raw.count("Origins")
            display_parts = []
            for i, p in enumerate(parts):
                display_parts.append(f"{p} Origins" if i < origins_count else p)

            bare_l1 = parts[0] if len(parts) >= 1 else ""
            bare_l2 = parts[1] if len(parts) >= 2 else ""
            bare_l3 = parts[2] if len(parts) >= 3 else ""

            level1 = display_parts[0] if len(display_parts) >= 1 else ""
            level2 = display_parts[1] if len(display_parts) >= 2 else ""
            level3 = display_parts[2] if len(display_parts) >= 3 else ""

            depth   = len(parts)
            # Trim administrative qualifiers (", not otherwise specified",
            # "(India)") down to the bare searchable word -- funding
            # descriptions say "Black", never "Black, not otherwise specified".
            raw_keyword = (bare_l3 or bare_l2 or bare_l1).strip().lower()
            keyword = extract_match_keyword(raw_keyword)
        else:
            # Fallback: read Level 1/2/3 columns directly if 'All Terms'
            # is blank/missing for this row. These already include the
            # "Origins" suffix, so no re-appending needed.
            level1 = safe_display(row.get(TAXONOMY_ENTRY1, ""))
            level2 = safe_display(row.get(TAXONOMY_ENTRY2, ""))
            level3 = safe_display(row.get(TAXONOMY_ENTRY3, ""))
            if not level1:
                continue
            depth   = 3 if level3 else (2 if level2 else 1)
            raw_keyword = clean_taxonomy_value(level3 or level2 or level1)
            keyword = extract_match_keyword(raw_keyword)

        if not level1:
            continue

        # Skip over-broad geographic keywords ("north american") that collide
        # with non-ethnic prose ("North American context/market") and never
        # resolve correctly on their own (see TAXONOMY_KEYWORD_STOPLIST).
        if keyword in TAXONOMY_KEYWORD_STOPLIST:
            continue

        entries.append({
            "keyword": keyword,
            "level1":  level1,
            "level2":  level2,
            "level3":  level3,
            "depth":   depth,
        })

    # Deepest first, then longest keyword first within the same depth
    entries.sort(key=lambda x: (-x["depth"], -len(x["keyword"])))
    return entries

# ==============================
# CASE-BY-CASE DETECTION LAYERS
# ==============================

def find_taxonomy_matches(text, taxonomy_entries):
    """
    Cases 1-3: direct taxonomy keyword matches, respecting negation
    and example-mention guards. Returns ALL valid matches found.
    The trailing s? allows plural forms ("Somalis", "Africans") to
    match the same singular taxonomy keyword without needing a
    separate entry for every pluralized variant.
    """
    matched = []
    for entry in taxonomy_entries:
        kw = entry["keyword"]
        if not kw:
            continue
        pattern = re.compile(r'\b' + re.escape(kw) + r's?\b', re.IGNORECASE)
        if not pattern.search(text):
            continue
        if is_negated(kw, text) or is_example_mention(kw, text):
            continue
        matched.append(entry)
    return matched

def find_pattern_match(text):
    """
    Case 4: structured/directional phrases not directly in taxonomy.
    Returns (l1, l2, l3, depth) or None.
    """
    for pattern, l1, l2, l3 in PATTERN_RULES:
        m = re.search(pattern, text, re.IGNORECASE)
        if not m:
            continue
        snippet = text[max(0, m.start()-80):m.start()]
        if matches_any(NEGATION_PHRASES, snippet):
            continue
        depth = 3 if l3 else (2 if l2 else 1)
        return (l1, l2, l3, depth)
    return None

def find_country_match(text):
    """
    Cases 6 & 7: nationality OR 'from <country>' phrasing.
    Returns (l1, l2, l3, depth) or None. The trailing s? allows plural
    forms ("Jamaicans", "Mexicans") to match without a separate entry
    for every pluralized variant.
    """
    for country, (l1, l2, l3) in COUNTRY_REGION_MAP.items():
        direct = re.search(r'\b' + re.escape(country) + r's?\b', text, re.IGNORECASE)
        from_phrase = re.search(r'\bfrom\s+' + re.escape(country) + r's?\b', text, re.IGNORECASE)
        m = direct or from_phrase
        if not m:
            continue
        snippet = text[max(0, m.start()-80):m.start()]
        if matches_any(NEGATION_PHRASES, snippet):
            continue
        depth = 3 if l3 else (2 if l2 else 1)
        return (l1, l2, l3, depth)
    return None

def is_bipoc_real_target(text):
    """Case 9 with context awareness: BIPOC mentioned as the actual
    target population vs. mentioned only as a general equity/mission
    statement (e.g. 'advancing equity', 'fostering BIPOC visibility'
    as one goal among several). We treat BIPOC as the target UNLESS
    it's wrapped in example-mention or override phrasing, or unless
    it appears alongside language suggesting it's an aspirational/
    values statement rather than a description of who is served.
    NOTE: This is ambiguous so: flagged
    for review rather than silently resolved either way.
    """
    found = False
    for kw_pattern in BIPOC_KEYWORDS:
        for m in re.finditer(kw_pattern, text, re.IGNORECASE):
            snippet = text[max(0, m.start()-80):m.start()]
            if matches_any(EXAMPLE_PHRASES, snippet):
                continue
            if matches_any(NEGATION_PHRASES, snippet):
                continue
            # "the BIPOC Grant / Media Lab / Program" names the funding stream,
            # not a served population -- skip this occurrence (a genuine "BIPOC
            # youth/women/artists" mention elsewhere still counts).
            after = text[m.end():m.end() + 40]
            if matches_any(BIPOC_PROGRAM_NAME_AFTER_PATTERNS, after):
                continue
            found = True
            break
        if found:
            break
    return found

def check_grassroots_case(text, has_ethnic_signal):
    """
    Case 11 (expanded):  'marginalized' / 'racialized' /
    'multicultural' / 'refugee' / 'immigrant' etc. never count as a signal
    on their own. Unlike negation/example-mention, this never suppresses
    silently either way -- caller flags the result whether a real signal
    is present alongside the buzzword or not.

    has_ethnic_signal is computed by the CALLER from the same candidate
    pool used for final classification in classify_row() -- this used
    to re-derive its own signal-detection pass via a separate
    has_real_ethnic_signal() helper, which silently missed BIPOC/POC
    (and could drift out of sync again the next time a new detection
    layer gets added to one path but not the other). Taking the
    already-computed signal as a parameter makes that class of bug
    structurally impossible -- there is only one detection pass now.
    """
    has_ambiguous = matches_any(AMBIGUOUS_EQUITY_WORDS, text)
    if not has_ambiguous:
        return None  # not relevant, no opinion
    if has_ethnic_signal:
        return "has_signal"  # real signal present -- classify normally, caller still flags
    return "no_signal"  # ambiguous word present but no real ethnic keyword

def check_org_name_lookup(text):
    """
    Case 10: known organization names. Only called as last resort.
    """
    for org_name, (l1, l2, l3) in ORG_NAME_ETHNICITY_MAP.items():
        if re.search(r'\b' + re.escape(org_name) + r'\b', text, re.IGNORECASE):
            return (l1, l2, l3)
    return None

def detect_broad_identity(text):
    return matches_any(BROAD_IDENTITY_KEYWORDS, text)

def context_is_overridden(text):
    return any(phrase_has_serving_context(p, text) for p in EXPANSION_PHRASES)

def context_is_historical(text):
    return any(phrase_has_serving_context(p, text) for p in HISTORICAL_PHRASES)

def context_is_aspirational(text):
    return matches_any(ASPIRATIONAL_PHRASES, text)

# =================================================
# CONTEXT SIGNAL LAYER
# Detects soft discourse signals in the full combined text.
# Results are annotations only — they MUST NOT affect classification decisions.
# =================================================

def extract_context_signals(text):
    return {
        "historical":   matches_any(HISTORICAL_PHRASES, text),
        "expansion":    matches_any(EXPANSION_PHRASES, text),
        "aspirational": matches_any(ASPIRATIONAL_PHRASES, text),
        "negation":     matches_any(ETHNIC_ANNOTATION_NEGATION_PHRASES, text),
        "example":      matches_any(EXAMPLE_PHRASES, text),
    }

def build_context_notes(signals, ethnic_term_present=False):
    """
    Production annotation notes.

    Only negation is surfaced — it directly affects how confident a reviewer
    should be in the ethnic evidence found (a negated term may or may not
    indicate the population served).

    Historical, expansion, aspirational, and example signals are omitted here
    because they are extremely common in nonprofit funding language and create
    alert fatigue when surfaced on every request.  They remain detectable via
    extract_context_signals() and are available in full via
    build_debug_context_notes() for debug overlay use.
    """
    notes = []
    # Anchor: only surface negation when an ethnic term was actually detected.
    if signals["negation"] and ethnic_term_present:
        notes.append(
            "Matched term appears near a negation word (e.g. \"not\", \"excluding\") "
            "- confirm the population is served, not excluded"
        )
    return notes

def build_debug_context_notes(signals):
    """
    Full context annotation notes including discourse-framing signals.
    For debug overlay use only — NOT emitted in production output.
    """
    notes = []
    if signals["historical"]:
        notes.append("Historical framing detected - may refer to past service scope only")
    if signals["expansion"]:
        notes.append("Scope expansion language detected - indicates broadened or non-exclusive targeting")
    if signals["aspirational"]:
        notes.append("Aspirational/future-oriented language detected - may not reflect current service population")
    if signals["negation"]:
        notes.append(
            "Matched term appears near a negation word (e.g. \"not\", \"excluding\") "
            "- confirm the population is served, not excluded"
        )
    if signals["example"]:
        notes.append("Example-based phrasing detected - referenced group may not be primary target")
    return notes

# =================================================
# CASE 13 — Potential Ethnocultural Organization Name
#
# Safety-net detector for unknown ethnocultural org names in the
# Funding Request Name column.  Only fires when Engine 1 produced zero
# candidates and BIPOC is absent — i.e. the row would otherwise fall
# through to General Population with no ethnic signal at all.
#
# NEVER classifies. NEVER overrides. Review flag only.
# =================================================

def looks_like_demonym(group_name):
    """Return True if group_name could plausibly be an ethnic/national identifier.
    Primary gate: any token in NON_ETHNIC_LEADING_WORDS → almost certainly not ethnic."""
    tokens = group_name.lower().split()
    return not any(t in NON_ETHNIC_LEADING_WORDS for t in tokens)

def is_known_taxonomy_keyword(group_name, taxonomy_entries):
    """Return True if group_name contains a recognized taxonomy keyword."""
    for entry in taxonomy_entries:
        kw = entry.get("keyword", "")
        if not kw:
            continue
        if re.search(r'\b' + re.escape(kw) + r's?\b', group_name, re.IGNORECASE):
            return True
    return False

def matches_pattern_rule(group_name):
    """Return True if group_name matches any PATTERN_RULES entry."""
    for pattern, *_ in PATTERN_RULES:
        if re.search(pattern, group_name, re.IGNORECASE):
            return True
    return False

def detect_ethnocultural_org_name(funding_request_name, candidates, bipoc_present, taxonomy_entries):
    """
    Case 13 — low-recall, high-precision safety net.

    Detects potential ethnocultural organization names in the Funding Request
    Name that Engine 1 did not already classify.

    Guard clause (Step 1): if ANY Engine 1 candidate exists or BIPOC is
    present, the row is already being handled — return None immediately.
    This ensures Case 13 only triggers as a true last resort.

    Negative filters (Step 6): even when the title matches the org-name
    pattern, do NOT flag if the extracted group token is:
      - a recognized taxonomy keyword
      - a known country/region/nationality in COUNTRY_REGION_MAP
      - matched by any PATTERN_RULES entry (directional ethnonyms etc.)
    """
    if candidates or bipoc_present:
        return None

    title = normalize_text(funding_request_name or "")
    if not title:
        return None

    for pattern in CASE_13_PATTERNS:
        match = re.search(pattern, title, re.IGNORECASE)
        if not match:
            continue

        group_name = match.group(1).strip()

        if is_known_taxonomy_keyword(group_name, taxonomy_entries):
            return None
        if group_name in COUNTRY_REGION_MAP:
            return None
        if matches_pattern_rule(group_name):
            return None
        if not looks_like_demonym(group_name):
            return None

        return "Note (low priority): Potential ethnocultural organization name detected - verify group identity manually"

    return None

# =================================================
# TEXT EXTRACTION — concatenate across all 4 columns
# =================================================

# Some compound identity phrases mean something specific that's
# different from their literal sub-words. "African Canadian" names the
# Black Canadian identity (-> Other Ethnic and Cultural Origins / Black,
# not otherwise specified) -- it is NOT the same as "African" alone
# (-> African Origins), but the bare word "african" matches that
# taxonomy entry on its own regardless of what follows it, so the
# phrase was getting silently reduced to the wrong, broader category.
# Rewriting the phrase to the word that already has a correct taxonomy
# entry means it flows through the EXISTING matching pipeline (negation
# guards, multi-group detection, depth resolution) automatically --
# no separate override system needed, and a second distinct group
# mentioned elsewhere in the same text (e.g. "African Canadian and
# Somali") still correctly produces Multiple instead of silently
# picking one and ignoring the other.
def apply_identity_phrase_rewrites(text):
    for pattern, replacement in IDENTITY_PHRASE_REWRITES:
        def _replace(m, text=text):
            preceding = text[max(0, m.start()-12):m.start()].strip(" -")
            if any(preceding.endswith(prefix) for prefix in DIRECTIONAL_AFRICAN_PREFIXES):
                return m.group(0)  # leave a regional phrase like "East African Canadian" untouched
            return replacement
        text = re.sub(pattern, _replace, text, flags=re.IGNORECASE)
    return text

def get_column_texts(row):
    texts = []
    for col in INPUT_COLS_PRIORITY:
        val = row.get(col, "")
        raw = normalize_text(val)
        raw = apply_identity_phrase_rewrites(raw)
        # Commented out for alias check implementation.
        #raw = apply_aliases(raw)
        texts.append(raw)
    return texts

def parse_raw_org_name(raw_name):
    """
    Plan.md Chunk G1 step 5 — strip the "Funding Request for ... NNNNN"
    wrapper and trailing FR-id off the RAW (pre-normalize, pre-identity-
    rewrite) funding-request account name, leaving just the org name text.
    Used by the generalizable silent-body name rule so identity-phrase
    rewrites (e.g. "african canadian" -> "africancanadian") never hide a
    term that should drive name-only classification.

    Tolerates both the canonical "NAME 12345" form and a display-truncated
    "NAME ...-12345" form (seen in some exported review sheets).
    Falls back to the raw text unchanged if it doesn't match either shape.
    """
    text = (raw_name or "").strip()
    if not text:
        return ""
    m = re.match(r'^funding request for\s+(.+)$', text, re.IGNORECASE)
    if m:
        text = m.group(1).strip()
    text = re.sub(r'\s*(\.\.\.)?[\s\-]*\d+\s*$', '', text).strip()
    return text

def _body_without_org_name(body_norm, raw_name):
    """Return the (already-normalized) body with the org's OWN name word
    tokens removed, so a self-reference echo ("Women Building Futures must
    replace servers", "The Ukrainian Shumka Dancers are undertaking...") is
    not mistaken for a served population by the cross-axis
    body_names_a_population() gate.

    Only meaningful in the silent-body name-rule branch, where the body has
    already been shown to carry no SERVED candidate on the classifying axis --
    so any body word that coincides with the org name is, by construction, the
    echo and safe to drop. A body that genuinely serves a population on
    ANOTHER axis (BCW 51622's "Haitian youth") keeps that term, since it is
    not part of the org name, so the gate still fires and blocks the borrow.
    """
    org = parse_raw_org_name(raw_name)
    if not org:
        return body_norm
    org_norm = apply_identity_phrase_rewrites(normalize_text(org))
    tokens = {w for w in re.findall(r'[a-z0-9]+', org_norm) if len(w) > 2}
    if not tokens:
        return body_norm
    pattern = r'\b(?:' + '|'.join(re.escape(t) for t in tokens) + r')\b'
    return re.sub(pattern, ' ', body_norm, flags=re.IGNORECASE)

def body_names_a_population(text):
    """
    Lightweight, taxonomy-independent check: does this (already normalized)
    body text name ANY identifiable served population at all -- ethnic,
    national, or gender/sexual-identity?

    Used to gate the generalizable silent-body name rule (Plan.md Chunk G1
    FIX): the name fallback on any ONE axis must fire ONLY when the body is
    truly administrative -- no served population named on ANY axis. e.g.
    BCW 51622's body serves "Haitian youth" -- a real (ethnic) served
    population that simply isn't gendered -- so the GENDER axis must stay
    General rather than borrowing "Women" from the org's own name; but a
    genuinely administrative body (bathroom-vanity replacement, ED-salary
    continuity, a website update) has no population signal on any axis, so
    the org name IS the best available evidence.

    Deliberately does NOT require taxonomy_entries (not available to the
    gender/sexual classifiers, which don't load the taxonomy sheet) --
    limited to the taxonomy-independent ethnic signal sources
    (COUNTRY_REGION_MAP, PATTERN_RULES, BROAD_IDENTITY_KEYWORDS, BIPOC) plus
    the gender/sexual term patterns (also plain regex, no taxonomy). This
    misses a body whose ONLY ethnic signal is a pure taxonomy keyword not
    covered by those sources (e.g. bare "Jewish"/"Cree") -- an accepted,
    narrow approximation given the cross-module constraint.
    """
    if not text:
        return False
    if is_bipoc_real_target(text):
        return True
    if matches_any(BROAD_IDENTITY_KEYWORDS, text):
        return True
    if find_pattern_match(text) is not None:
        return True
    if find_country_match(text) is not None:
        return True
    if any(re.search(p, text, re.IGNORECASE) for p, _, _ in GENDER_TERM_PATTERNS):
        return True
    if matches_any(SEXUAL_ORIENTATION_PATTERNS, text):
        return True
    return False

def get_body_and_name_texts(row):
    """Return (body_text, name_text), each normalized + identity-rewritten
    exactly like get_column_texts, so callers can require a served-population
    signal in the body and treat a name-only signal as unclassified."""
    def _prep(val):
        return apply_identity_phrase_rewrites(normalize_text(val))
    body = " ".join(t for t in (_prep(row.get(c, "")) for c in BODY_COLS) if t.strip())
    name = _prep(row.get(NAME_COL, ""))
    return body, name

# ==========================
# MAIN CLASSIFICATION LOGIC
# =============================

# =====
# MAIN
# =====
def main():
    # Deferred import avoids circular dependency: classify_pipeline imports from this module.
    from classify_pipeline import classify_row as pipeline_classify_row

    start_time = time.time()

    ds = bootstrap.dataset()
    print(f"[dataset] active: {ds.name}  ->  reads {ds.raw_file.name}, writes {ds.output_file.name}")

    taxonomy_filepath = ds.taxonomy_file
    funding_filepath = ds.raw_file

    print(f"Loading taxonomy from: {taxonomy_filepath}")
    try:
        tax_df = pd.read_excel(taxonomy_filepath, sheet_name=TAXONOMY_SHEET, dtype=str)
    except Exception as e:
        print(f"Error loading taxonomy sheet '{TAXONOMY_SHEET}' from '{taxonomy_filepath}': {e}")
        sys.exit(1)

    print(f"Loading funding requests from: {funding_filepath}")
    if not funding_filepath.exists():
        print(f"Error: data workbook not found at '{funding_filepath}'.")
        sys.exit(1)
    try:
        data_df = pd.read_excel(funding_filepath, sheet_name=DATA_SHEET, dtype=str)
    except Exception as e:
        print(f"Error loading data sheet '{DATA_SHEET}' from '{funding_filepath}': {e}")
        sys.exit(1)
 
    print(f"Taxonomy rows: {len(tax_df)} | Data rows: {len(data_df)}")
 
    taxonomy_entries = build_taxonomy(tax_df)
    print(f"Taxonomy entries parsed: {len(taxonomy_entries)}")

    # Semantic Fallback - Level 2 Engine
    semantic_entries = None
    semantic_embeddings = None

    if SEMANTIC_AVAILABLE:
        scope_notes = semantic_fallback.build_scope_note_map(
            tax_df, TAXONOMY_ENTRY1, TAXONOMY_SCOPE_NOTES, safe_display)
        semantic_entries, semantic_texts = semantic_fallback.build_candidate_texts(
            taxonomy_entries, scope_notes)
        semantic_embeddings = semantic_fallback.get_taxonomy_embeddings(semantic_texts)
    else:
        print("semantic_fallback not available (sentence-transformers not installed) — skipping semantic suggestions.")
 
    for col in [OUTPUT_ETHNIC1, OUTPUT_ETHNIC2, OUTPUT_ETHNIC3, OUTPUT_FLAG, OUTPUT_SEMANTIC]:
        if col not in data_df.columns:
            data_df[col] = ""
    # Count how many rows fall into each outcome bucket, for summary stats at the end. Note that these are not mutually exclusive categories (e.g. a row with a pattern match that's also flagged for aspirational language would count in both "pattern" and "flagged"), but they give a general sense of how many hits came from each detection method and how many were flagged for review.
    # defaultdict so the diagnostic counters incremented below
    # (stats["pattern"], stats["country"], stats["org_lookup"],
    # stats["semantic_suggested"], stats["flagged"]) auto-initialize — these
    # keys were never in the seeded dict, which KeyError'd the full main() run
    # (the unit tests exercise classify_row directly and never hit this path).
    stats = defaultdict(int, {"-Ethnic → 3-level": 0, "-Ethnic → 2-level": 0, "-Ethnic → 1-level": 0, "-Ethnic → Multiple": 0,
              "-Ethnic → Other": 0, "-Ethnic → General": 0, "-Flagged → Ethnic": 0, "-Pattern → Ethnic": 0,
              "-Country → Ethnic": 0, "-Organization Name → Ethnic": 0, "-Grassroots Filtered → Ethnic": 0, "-Semantic Suggested → Ethnic": 0})
    
    print(f"\n --- Classification Categories ---\n")
    
    for idx, row in data_df.iterrows():
        # Route through the canonical refactored pipeline (classify_pipeline → resolver)
        e1, e2, e3, flag = pipeline_classify_row(row, taxonomy_entries)
        data_df.at[idx, OUTPUT_ETHNIC1] = e1
        data_df.at[idx, OUTPUT_ETHNIC2] = e2
        data_df.at[idx, OUTPUT_ETHNIC3] = e3
        data_df.at[idx, OUTPUT_FLAG] = flag

        # Engine 2 - This is our Semantic fallback layer, which only triggers if Engine 1 returns General Population (i.e. no
        if SEMANTIC_AVAILABLE and e1 == GENERAL_POP:
            combined_text = " ".join(t for t in get_column_texts(row) if t.strip())
            suggestion = semantic_fallback.find_semantic_suggestion(
                combined_text, semantic_entries, semantic_embeddings)
            if suggestion:
                sl1, sl2, sl3, score, margin = suggestion
                parts = [p for p in [sl1, sl2, sl3] if p]
                data_df.at[idx, OUTPUT_SEMANTIC] = f"{' / '.join(parts)} (similarity: {score:.2f}, margin: {margin:.2f})"
                stats["semantic_suggested"] += 1
 
        if e1 == MULTIPLE_ETHNIC:
            stats["-Ethnic → Multiple"] += 1
        elif e1 == OTHER_ETHNIC:
            stats["-Ethnic → Other"] += 1
        elif e1 == GENERAL_POP:
            stats["-Ethnic → General"] += 1
        elif e3:
            stats["-Ethnic → 3-level"] += 1
        elif e2:
            stats["-Ethnic → 2-level"] += 1
        else:
            stats["-Ethnic → 1-level"] += 1
 
        if "pattern rule" in flag.lower():
            stats["pattern"] += 1
        if "country" in flag.lower():
            stats["country"] += 1
        if "organization name" in flag.lower():
            stats["org_lookup"] += 1
        # if "ambiguous equity" in flag.lower():
        #     stats["grassroots_filtered"] += 1
        if flag:
            stats["flagged"] += 1
 
    wb = load_workbook(funding_filepath)
    ws = wb[DATA_SHEET]
    headers = {cell.value: cell.column for cell in ws[1]}

    for col_name in [OUTPUT_ETHNIC1, OUTPUT_ETHNIC2, OUTPUT_ETHNIC3, OUTPUT_FLAG, OUTPUT_SEMANTIC]:
        if col_name not in headers:
            new_col = ws.max_column + 1
            ws.cell(row=1, column=new_col, value=col_name)
            headers[col_name] = new_col

    for i, (idx, row) in enumerate(data_df.iterrows(), start=2):
        ws.cell(row=i, column=headers[OUTPUT_ETHNIC1], value=data_df.at[idx, OUTPUT_ETHNIC1])
        ws.cell(row=i, column=headers[OUTPUT_ETHNIC2], value=data_df.at[idx, OUTPUT_ETHNIC2])
        ws.cell(row=i, column=headers[OUTPUT_ETHNIC3], value=data_df.at[idx, OUTPUT_ETHNIC3])
        ws.cell(row=i, column=headers[OUTPUT_FLAG],    value=data_df.at[idx, OUTPUT_FLAG])
        ws.cell(row=i, column=headers[OUTPUT_SEMANTIC], value=data_df.at[idx, OUTPUT_SEMANTIC]) # Writes "" for every row that didn't receive a suggestion (General pop.)
    wb.save(ds.output_file)

    print("\nResults:")
    for k, v in stats.items():
        print(f"  {k}: {v}")

    print(f"\nOutput written to: {ds.output_file}")
    elapsed = time.time() - start_time
    print(f"Ethnic classification completed in {elapsed:.1f} seconds.")
 
if __name__ == "__main__":
    main()
