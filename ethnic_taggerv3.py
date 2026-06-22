import sys
import re
import pandas as pd
import time

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

"""
ethnic_tagger_v3.py
--------------------
Rule-based ethnic origin classifier for ECF Discretionary
Funding Requests.

Reads:
    "Taxonomy - Definitions.xlsx" -> sheet "Ethnic and Cultural Origins"
    "Discretionary FR working -2025 (Oreva).xlsx" -> sheet "Discretionary Funding Requests"

Aggregates matches across all 4 input columns (Final_Project_Description,
Final_Summary_Description, Purpose, Funding Request Name), then resolves
using the case hierarchy below.

To Run:
    python ethnic_taggerv3.py "C:\\Users\\oadode\\OneDrive - Edmonton Community Foundation\\Desktop\\Discretionary FR Scripting\\ECF-Discretionary-FR-Scripting\\Taxonomy - Definitions.xlsx" "C:\\Users\\oadode\\OneDrive - Edmonton Community Foundation\\Desktop\\Discretionary FR Scripting\\ECF-Discretionary-FR-Scripting\\FR testing.xlsx"

Case coverage (see README):
    1.  Exact match (deepest taxonomy term)
    2.  Level 2 match (subregion)
    3.  Level 1 match (broad category)
    4.  Structured phrase ("South African" etc.) -> pattern rules
    5.  Country/nationality IN taxonomy -> handled by case 1-3
    6.  Country/nationality NOT in taxonomy -> country map
    7.  "from <country>" phrasing -> country map (phrase variant)
    8.  Multiple groups -> Multiple Ethnic and Cultural Origins
    9.  BIPOC (context-aware) -> Multiple Ethnic and Cultural Origins
    9b. Broad identity labels (Black, Arab, Jewish, etc.) -> Other Ethnic
        and Cultural Origins
    10. Known organization name lookup (Bent Arrow, Treaty 6, etc.)
        -> only consulted if result would otherwise be General Population
    11. "Grassroots" -> ethnic-indicating ONLY if paired with another
        ethnic keyword; otherwise ignored entirely
    12. General Population (fallback / no signal)

Also handles:
    - Context override ("beyond its original focus on X")
        - Historical reference ("historically served X")
        - Expansion Phrases ("")
    - Negation ("not targeting X", "does not serve X")
    - Aspirational language ("plans to expand to X") -> flagged for review
    - Example mentions ("such as X", "including X") -> not classified
    - Strict word-boundary matching (no "blacksmith" -> "black")
"""

# Semantic-similarity fallback (see semantic_fallback.py).
try:
    import semantic_fallback
    SEMANTIC_AVAILABLE = True
except ImportError:
    SEMANTIC_AVAILABLE = False

# =============
# CONFIGURATION 
# =============

TAXONOMY_SHEET = "Ethnic and Cultural Origins"
DATA_SHEET = "Discretionary Funding Requests"

TAXONOMY_ENTRY1 = "Ethnic and Cultural Origins Level 1"
TAXONOMY_ENTRY2 = "Ethnic and Cultural Origins Level 2"
TAXONOMY_ENTRY3 = "Ethnic and Cultural Origins Level 3"
TAXONOMY_ALL_TERMS = "All Terms"
TAXONOMY_SCOPE_NOTES = "Definitions and Scope Notes"

# Search priority order. All 4 are concatenated, but this order
# determines which match "wins" on ties (first column listed wins)
INPUT_COLS_PRIORITY = [
    "Final_Project_Description",
    "Final_Summary_Description",
    "Purpose",
    "Funding Request Name",
]

OUTPUT_ETHNIC1 = "Ethnic 1 - FR6"
OUTPUT_ETHNIC2 = "Ethnic 2 - FR7"
OUTPUT_ETHNIC3 = "Ethnic 3 - FR8"
OUTPUT_FLAG = "Classification Flag"
OUTPUT_SEMANTIC = "Semantic Suggestion (REVIEW)"

MULTIPLE_ETHNIC = "Multiple Ethnic and Cultural Origins"
OTHER_ETHNIC = "Other Ethnic and Cultural Origins"
GENERAL_POP = "General Population (No specific ethnic and cultural origin group served)"

# =======================
# CASE 9 — BIPOC keywords 
# =======================
BIPOC_KEYWORDS = [
    r"\bbipoc\b",
    r"\bqtbipoc\b",
    r"\bpoc\b",
    r"\bpeople of colou?r\b",
    r"\bblack african\b",
    #r"\bracialized\b", # Change to flag if "racialized" is detected
]

# Words that, on their own, should NOT trigger BIPOC/Multiple classification
# unless paired with an actual ethnic 'hint'. Handled separately from
# BIPOC_KEYWORDS because the rule is different (Case 11).
# Always flagged for review either way now (signal or no signal) --
# see check_grassroots_case().

AMBIGUOUS_EQUITY_WORDS = [
    r"\bmarginalized\b",
    r"\bgrassroots\b",
    r"\bethnocultural\b",
    r"\bracialized\b",
    r"\bunderrepresented\b",
    r"\bmulticultural\b",
    r"\bdiverse\b",
    r"\brefugee\b",
    r"\bimmigrant\b",
    r"\bfrancophone\b",
    r"\bnewcomer\b",
    r"\bculturally\b",
]

# ============================================================
# CASE 9b — Broad identity labels (not in taxonomy directly)
# Black/Jewish/Arab removed -- these are real Level 2 entries under
# "Other Ethnic and Cultural Origins" in the taxonomy, matched through
# find_taxonomy_matches() instead of here now.
# ============================================================
BROAD_IDENTITY_KEYWORDS = [
    r"\bhispanic\b",
    r"\blatino\b",
    r"\blatina\b",
    r"\blatinx\b",
    r"\bmixed heritage\b",
    r"\bmixed race\b",
    r"\bmultiracial\b",
    r"\bmulti[\-\s]ethnic\b",
]

# Afro-Caribbean / Afro-Latino name TWO distinct Level 1 groups at
# once -- always Multiple, regardless of anything else in the text.
ALWAYS_MULTIPLE_COMPOUNDS = {
    r"\bafro[\-\u2010\u2011\u2012\u2013\u2014\s]*caribbean\b": (
        ("African Origins", "", ""),
        ("Caribbean Origins", "", ""),
    ),
    r"\bafro[\-\u2010\u2011\u2012\u2013\u2014\s]*latin(o|a|x)?\b": (
        ("African Origins", "", ""),
        ("Latin American Origins", "", ""),
    ),
}

# =========================================================
# CASE 10 — Known organization name -> ethnicity lookup
# Only consulted as a LAST RESORT if classification would
# otherwise be General Population.
# ========================================================
ORG_NAME_ETHNICITY_MAP = {
    "bent arrow": ("North American Indigenous Origins", "", ""),
    "treaty 6": ("North American Indigenous Origins", "", ""), # Flag for review, as "Treaty 6" could refer to the geographic region (which would be L1 or L2) rather than the organization (which would be Case 10). Only trigger if "Treaty 6" appears in the funding request name or purpose, not just the description.
    # Add more known organization name -> ethnicity mappings here as identified
}

# ==================================================
# CASE 4 — Structured / directional phrase patterns
# Applied only if no direct taxonomy match found.
# ==================================================
PATTERN_RULES = [
    (r"\bnorth[\s\-]?african\b", "African Origins", "North African Origins", ""),
    (r"\bsouth[\s\-]?african\b", "African Origins", "Southern and East African Origins", ""),
    (r"\beast[\s\-]?african\b", "African Origins", "Southern and East African Origins", ""),
    (r"\bwest[\s\-]?african\b", "African Origins", "Central and West African Origins", ""),
    (r"\bcentral[\s\-]?african\b", "African Origins", "Central and West African Origins", ""),
    (r"\bsub[\s\-]?saharan\b", "African Origins", "Central and West African Origins", ""),
    (r"\bsoutheast[\s\-]?asian\b", "Asian Origins", "East and Southeast Asian Origins", ""),
    (r"\bsouth[\s\-]?asian\b", "Asian Origins", "South Asian Origins", ""),
    (r"\beast[\s\-]?asian\b", "Asian Origins", "East and Southeast Asian Origins", ""),
    (r"\bmiddle[\s\-]?eastern\b", "Asian Origins", "West and Central Asian and Middle Eastern Origins", ""),
    (r"\bfirst nations\b", "North American Indigenous Origins", "", ""),
    (r"\bmetis\b", "North American Indigenous Origins", "", ""),
    (r"\binuit\b", "North American Indigenous Origins", "", ""),
    (r"\baboriginal\b","North American Indigenous Origins", "", ""),
    (r"\bindigenous canadian\b", "North American Indigenous Origins", "", ""),
    (r"\btreaty 6\b", "North American Indigenous Origins", "", ""),
    (r"\bnorthern european\b", "European Origins", "Northern European Origins", ""),
    (r"\bsouthern european\b", "European Origins", "Southern European Origins", ""),
    (r"\beast(ern)? european\b", "European Origins", "Eastern European Origins", ""),
    (r"\bwest(ern)? european\b", "European Origins", "Western European Origins", ""),
]

# ============================================================
# CASES 6 & 7 — Country/nationality NOT in taxonomy
# Covers both "Jamaican" (nationality) and "from Jamaica" (Case 7)
# ============================================================
COUNTRY_REGION_MAP = {
    "jamaican": ("Caribbean Origins", "", ""),
    "jamaica": ("Caribbean Origins", "", ""),
    "trinidadian": ("Caribbean Origins", "", ""),
    "trinidad": ("Caribbean Origins", "", ""),
    "barbadian": ("Caribbean Origins", "", ""),
    "barbados": ("Caribbean Origins", "", ""),
    "haitian": ("Caribbean Origins", "", ""),
    "haiti": ("Caribbean Origins", "", ""),
    "guyanese": ("Caribbean Origins", "", ""),
    "guyana": ("Caribbean Origins", "", ""),
    "brazilian": ("Latin American Origins", "", ""),
    "brazil": ("Latin American Origins", "", ""),
    "colombian": ("Latin American Origins", "", ""),
    "colombia": ("Latin American Origins", "", ""),
    "mexican": ("Latin American Origins", "", ""),
    "mexico": ("Latin American Origins", "", ""),
    "salvadoran": ("Latin American Origins", "", ""),
    "el salvador": ("Latin American Origins", "", ""),
    "guatemalan": ("Latin American Origins", "", ""),
    "guatemala": ("Latin American Origins", "", ""),
    "peruvian": ("Latin American Origins", "", ""),
    "peru": ("Latin American Origins", "", ""),
    "venezuelan": ("Latin American Origins", "", ""),
    "venezuela": ("Latin American Origins", "", ""),
    "indian": ("Asian Origins", "South Asian Origins", "Indian (India)"),
    "india": ("Asian Origins", "South Asian Origins", "Indian (India)"),
    "kerala": ("Asian Origins", "South Asian Origins", "Indian (India)"), # state in India, often named directly e.g. "Kerala Cultural Association"
    "cameroonian": ("African Origins", "Central and West African Origins", ""), # no L3 entry for Cameroon -- falls to L2
    "cameroon": ("African Origins", "Central and West African Origins", ""),
}

# ============================================================
# Context-override / historical / negation / aspirational / example
# phrase banks — all generic, applied to ANY keyword (not group-specific)
# ============================================================
EXPANSION_PHRASES = [
    r"beyond (its|their|our|the) (original|previous|former|initial|traditional|historic(al)?)",
    r"expanding beyond",
    r"expansion",
    r"not (only|exclusively|solely|limited to|just|restricted to)",
    r"no longer (limited|restricted|focused|exclusively)",
    r"open(ing)? (up )?to (all|broader|wider|diverse|other)",
    r"(increasingly|more) diverse",
    r"welcom(es?|ing) (all|everyone|anyone|diverse)",
    r"regardless of (ethnic|cultural|racial|background)",
    r"all (cultural|ethnic|racial)? backgrounds",
    r"without regard to",
    r"irrespective of",
    r"inclusive of all",
]

HISTORICAL_PHRASES = [
    r"historic(al(ly)?)?",
    r"former(ly)?",
    r"previous(ly)?",
    r"original(ly)?",
    r"(in|during) the past",
    r"used to (serve|focus|target|support)",
    r"once (served|focused|targeted|supported)",
    r"(its|their|our) roots (in|with)",
    r"founded (to serve|for|by)",
    r"(was|were) (established|created|founded) (for|to serve)",
]

NEGATION_PHRASES = [
    r"not (target(ing)?|serv(ing|e)|focus(ing|ed)|for|limited to|exclusively)",
    r"does not (target|serve|focus|support|cater)",
    r"do not (target|serve|focus|support|cater)",
    r"no longer (target(ing)?|serv(ing|e)|focus(ing|ed))",
    r"exclud(es?|ing)",
    r"except(ing)?",
    r"other than",
    r"outside of",
    r"not the (primary|main|sole|only) (focus|target|group|population)",
]

ASPIRATIONAL_PHRASES = [
    r"hop(es?|ing) to (serve|reach|support|engage|include|target)",
    r"plan(s|ning) to (serve|reach|support|engage|include|target)",
    r"aim(s|ing) to (serve|reach|support|engage|include|target)",
    r"intend(s|ing) to",
    r"will (eventually|soon|begin to|start to) (serve|reach|support)",
    r"goal(s)? (is|are|of|to) (reach(ing)?|serv(ing|e)|includ(ing|e))",
    r"aspir(es?|ing) to",
    r"seek(s|ing) to (expand|reach|grow|include)",
    r"in the future",
    r"(future|upcoming) (focus|programming|initiative)",
]

EXAMPLE_PHRASES = [
    r"such as",
    r"for example",
    r"e\.g\.?",
    r"i\.e\.?",
    r"includ(ing|e) (communities|groups|populations|organizations|people)? ?(such as|like)",
    r"like (the )?following",
    r"among (others|other groups|other communities)",
    r"(and|or) (other|similar) (communities|groups|populations)",
    r"compar(ed|ing) to",
    r"as (opposed|compared) to",
]

# Words suggesting an ethnic term names a CONSULTED PARTY (expert/
# advisor role) rather than the population served, e.g. "...consult
# wildlife biologists, conservation experts, and indigenous knowledge
# holders". Never suppresses a match -- only adds a flag, since there's
# no reliable way to confirm this either way.
EXPERT_ROLE_PHRASES = [
    r"\bexperts?\b",
    r"\bspecialists?\b",
    r"\bconsultants?\b",
    r"\badvisors?\b",
    r"\bpractitioners?\b",
    r"\bknowledge holders?\b",
    r"\bstakeholders?\b",
    r"\bbiologists?\b",
]

"""
-- Not Necessary to be handled right now for processing sakes.

# Common typos / nationality-vs-canonical-term variants.
KEYWORD_ALIASES = {
    "somalian": "somali",
    "ethopian": "ethiopian",
    "ethipian": "ethiopian",
    "rwandese": "rwandan",
    "congolaise": "congolese",
    "congolais": "congolese",
    "mozambiquean": "mozambican",
    "tanzanean": "tanzanian",
    "ugandese": "ugandan",
    "burundaise": "burundian",
    "filippino": "filipino",
    "phillipine": "filipino",
    "philippine": "filipino",
    "viet": "vietnamese",
    "indo-canadian": "south asian",
    "indo canadian": "south asian",
    "south-asian": "south asian",
    "middle eastern": "west and central asian and middle eastern",
    "middle-eastern": "west and central asian and middle eastern",
}
"""

# =======
# HELPERS
# =======

def clean_taxonomy_value(val):
    """Lowercase, strip, and remove the literal word 'origins'. Treats NaN/None as
    empty. Pandas reads blank Excel cells as float('nan'), so an explicit pd.isna() check is required here."""
    if val is None or (isinstance(val, float) and pd.isna(val)) or pd.isna(val):
        return ""
    val = str(val).strip()
    if val.lower() == "nan":
        return ""
    val = val.lower()
    val = val.replace("origins", "")
    return val.strip()

def safe_display(val):
    """Return the original display value for level1/level2/level3,
    treating NaN as empty string instead of the literal 'nan'."""
    if val is None or pd.isna(val):
        return ""
    val = str(val).strip()
    return "" if val.lower() == "nan" else val

def extract_match_keyword(raw_keyword):
    """Trim a taxonomy label down to the bare word(s) worth searching
    for in funding-request text: "black, not otherwise specified" ->
    "black", "indian (india)" -> "indian"."""
    keyword = raw_keyword.split(",")[0]
    keyword = re.sub(r"\s*\([^)]*\)\s*$", "", keyword)
    return keyword.strip()

def normalize_text(text):
    """Lowercase, strip punctuation, collapse whitespace. Used on the
    funding-request free text columns (not the taxonomy sheet, which
    uses clean_taxonomy_value instead)."""
    if pd.isna(text) or not isinstance(text, str):
        return ""
    text = text.lower()
    text = re.sub(r'[^\w\s]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

# UNCOMMENT : For implementation of alias check
# ------
# def apply_aliases(text):
#     for alias, canonical in KEYWORD_ALIASES.items():
#         text = re.sub(r'\b' + re.escape(alias) + r'\b', canonical, text)
#     return text

def matches_any(patterns, text):
    return any(re.search(p, text, re.IGNORECASE) for p in patterns)

def keyword_context_window(keyword, text, window=80):
    """Return the text immediately preceding the first occurrence of
    keyword, used to check for negation/example/override phrases that
    appear right before it."""
    pattern = re.compile(r'\b' + re.escape(keyword) + r'\b', re.IGNORECASE)
    m = pattern.search(text)
    if not m:
        return None
    start = max(0, m.start() - window)
    return text[start:m.start()]

def is_negated(keyword, text):
    snippet = keyword_context_window(keyword, text)
    return snippet is not None and matches_any(NEGATION_PHRASES, snippet)

def is_example_mention(keyword, text):
    snippet = keyword_context_window(keyword, text)
    return snippet is not None and matches_any(EXAMPLE_PHRASES, snippet)

SERVING_CONTEXT_WORDS = [
    r"\bserve(s|d)?\b",
    r"\bserving\b",
    r"\bpopulation\b",
    r"\bcommunit(y|ies)\b",
    r"\bdemographic(s)?\b",
    r"\bfocus(ed|es)?\b",
    r"\btarget(ed|ing)?\b",
    r"\breach\b",
    r"\bbeneficiar(y|ies)\b",
    r"\bclientele\b",
    r"\bmembership\b",
    r"\bmembers\b",
    r"\baudience\b",
    r"\bclients\b",
    r"\bresidents\b",
    r"\bgroups\b",
]

def phrase_has_serving_context(pattern, text, window=100):
    for m in re.finditer(pattern, text, re.IGNORECASE):
        start = max(0, m.start() - window)
        end = min(len(text), m.end() + window)
        if matches_any(SERVING_CONTEXT_WORDS, text[start:end]):
            return True
    return False

# ============================================================
# TAXONOMY BUILDER (Case 1-3 source data)
# Sort: deepest first, then LONGEST keyword first within same depth. Prevents "African"
# from matching before "Southern and East African".
# ============================================================

# ============================================================
# TAXONOMY BUILDER (Case 1-3 source data)
#
# Pulled from Column D "All Terms", a flat concatenation of
# Level 1 + Level 2 + Level 3 with the literal word "Origins" acting
# as the delimiter after L1 and L2 (no separator after L3,
# since specific terms like "Somali" don't carry an "Origins" suffix).
#
# Example: "African OriginsSouthern and East African OriginsSomali"
#   -> split on "Origins" -> ["African", "Southern and East African", "Somali"]
#
# Falls back to reading Level 1/2/3 columns directly if All Terms is
# blank or missing for a given row, for resilience against a broken
# formula cell.
#
# Sort: deepest first, then LONGEST keyword first within the same
# depth. This is the fix from the production skeleton — prevents
# "African" from matching before "Southern and East African".
# ============================================================

def parse_all_terms(all_terms_value):
    """
    Split the 'All Terms' cell on the literal word 'Origins' into
    its component level strings, stripped of whitespace. Trailing chars removed.
    """
    if pd.isna(all_terms_value) or not isinstance(all_terms_value, str):
        return []
    parts = all_terms_value.split("Origins")
    parts = [p.strip() for p in parts if p.strip()]
    return parts

def build_taxonomy(tax_df):
    entries = []
    for _, row in tax_df.iterrows():
        all_terms_raw = row.get(TAXONOMY_ALL_TERMS, "")
        parts = parse_all_terms(all_terms_raw)

        if parts:
            # Primary path: parsed from All Terms column.
            # The split strips the literal word "Origins" from every
            # fragment, but in the real taxonomy, Level 1 and Level 2
            # display values usually INCLUDE the "Origins" suffix
            # (e.g. "African Origins", "Southern and East African Origins")
            # -- except entries like "Black, not otherwise specified" and
            # "Jewish" under "Other Ethnic and Cultural Origins", which
            # only have "Origins" on Level 1. Count actual occurrences of
            # "Origins" in the source string so we know how many leading
            # parts really need it re-appended, instead of assuming L1+L2
            # always do.
            origins_count = all_terms_raw.count("Origins")
            display_parts = []
            for i, p in enumerate(parts):
                display_parts.append(f"{p} Origins" if i < origins_count else p)

            bare_l1 = parts[0] if len(parts) >= 1 else ""
            bare_l2 = parts[1] if len(parts) >= 2 else ""
            bare_l3 = parts[2] if len(parts) >= 3 else ""

            level1 = display_parts[0] if len(display_parts) >= 1 else ""
            level2 = display_parts[1] if len(display_parts) >= 2 else ""
            level3 = display_parts[2] if len(display_parts) >= 3 else ""

            depth   = len(parts)
            # Trim administrative qualifiers (", not otherwise specified",
            # "(India)") down to the bare searchable word -- funding
            # descriptions say "Black", never "Black, not otherwise specified".
            raw_keyword = (bare_l3 or bare_l2 or bare_l1).strip().lower()
            keyword = extract_match_keyword(raw_keyword)
        else:
            # Fallback: read Level 1/2/3 columns directly if 'All Terms'
            # is blank/missing for this row. These already include the
            # "Origins" suffix, so no re-appending needed.
            level1 = safe_display(row.get(TAXONOMY_ENTRY1, ""))
            level2 = safe_display(row.get(TAXONOMY_ENTRY2, ""))
            level3 = safe_display(row.get(TAXONOMY_ENTRY3, ""))
            if not level1:
                continue
            depth   = 3 if level3 else (2 if level2 else 1)
            raw_keyword = clean_taxonomy_value(level3 or level2 or level1)
            keyword = extract_match_keyword(raw_keyword)

        if not level1:
            continue

        entries.append({
            "keyword": keyword,
            "level1":  level1,
            "level2":  level2,
            "level3":  level3,
            "depth":   depth,
        })

    # Deepest first, then longest keyword first within the same depth
    entries.sort(key=lambda x: (-x["depth"], -len(x["keyword"])))
    return entries

# ==============================
# CASE-BY-CASE DETECTION LAYERS
# ==============================

def find_taxonomy_matches(text, taxonomy_entries):
    """
    Cases 1-3: direct taxonomy keyword matches, respecting negation
    and example-mention guards. Returns ALL valid matches found.
    """
    matched = []
    for entry in taxonomy_entries:
        kw = entry["keyword"]
        if not kw:
            continue
        pattern = re.compile(r'\b' + re.escape(kw) + r'\b', re.IGNORECASE)
        if not pattern.search(text):
            continue
        if is_negated(kw, text) or is_example_mention(kw, text):
            continue
        matched.append(entry)
    return matched

def find_pattern_match(text):
    """
    Case 4: structured/directional phrases not directly in taxonomy.
    Returns (l1, l2, l3, depth) or None.
    """
    for pattern, l1, l2, l3 in PATTERN_RULES:
        m = re.search(pattern, text, re.IGNORECASE)
        if not m:
            continue
        snippet = text[max(0, m.start()-80):m.start()]
        if matches_any(NEGATION_PHRASES, snippet):
            continue
        depth = 3 if l3 else (2 if l2 else 1)
        return (l1, l2, l3, depth)
    return None

def find_country_match(text):
    """
    Cases 6 & 7: nationality OR 'from <country>' phrasing.
    Returns (l1, l2, l3, depth) or None.
    """
    for country, (l1, l2, l3) in COUNTRY_REGION_MAP.items():
        direct = re.search(r'\b' + re.escape(country) + r'\b', text, re.IGNORECASE)
        from_phrase = re.search(r'\bfrom\s+' + re.escape(country) + r'\b', text, re.IGNORECASE)
        m = direct or from_phrase
        if not m:
            continue
        snippet = text[max(0, m.start()-80):m.start()]
        if matches_any(NEGATION_PHRASES, snippet):
            continue
        depth = 3 if l3 else (2 if l2 else 1)
        return (l1, l2, l3, depth)
    return None

def has_real_ethnic_signal(text, taxonomy_entries):
    """
    Used by Case 11 (grassroots) to check whether an ACTUAL ethnic
    keyword exists elsewhere in the text, independent of the ambiguous
    equity words.
    """
    # This will later be revised, as grassroots can benfit any ethnic group
    if find_taxonomy_matches(text, taxonomy_entries):
        return True
    if find_pattern_match(text):
        return True
    if find_country_match(text):
        return True
    for compound_pattern in ALWAYS_MULTIPLE_COMPOUNDS:
        if re.search(compound_pattern, text, re.IGNORECASE):
            return True
    if detect_broad_identity(text):
        return True
    return False

def is_bipoc_real_target(text):
    """Case 9 with context awareness: BIPOC mentioned as the actual
    target population vs. mentioned only as a general equity/mission
    statement (e.g. 'advancing equity', 'fostering BIPOC visibility'
    as one goal among several). We treat BIPOC as the target UNLESS
    it's wrapped in example-mention or override phrasing, or unless
    it appears alongside language suggesting it's an aspirational/
    values statement rather than a description of who is served.
    NOTE: This is ambiguous so: flagged
    for review rather than silently resolved either way.
    """
    found = False
    for kw_pattern in BIPOC_KEYWORDS:
        m = re.search(kw_pattern, text, re.IGNORECASE)
        if m:
            snippet = text[max(0, m.start()-80):m.start()]
            if matches_any(EXAMPLE_PHRASES, snippet):
                continue
            if matches_any(NEGATION_PHRASES, snippet):
                continue
            found = True
            break
    return found

def check_grassroots_case(text, taxonomy_entries):
    """
    Case 11 (expanded): 'grassroots' / 'marginalized' / 'ethnocultural' /
    'multicultural' / 'refugee' / 'immigrant' etc. never count as a signal
    on their own. Unlike negation/example-mention, this never suppresses
    silently either way -- caller flags the result whether a real signal
    is present alongside the buzzword or not.
    """
    has_ambiguous = matches_any(AMBIGUOUS_EQUITY_WORDS, text)
    if not has_ambiguous:
        return None  # not relevant, no opinion
    if has_real_ethnic_signal(text, taxonomy_entries):
        return "has_signal"  # real signal present -- classify normally, caller still flags
    return "no_signal"  # ambiguous word present but no real ethnic keyword

def check_org_name_lookup(text):
    """
    Case 10: known organization names. Only called as last resort.
    """
    for org_name, (l1, l2, l3) in ORG_NAME_ETHNICITY_MAP.items():
        if re.search(r'\b' + re.escape(org_name) + r'\b', text, re.IGNORECASE):
            return (l1, l2, l3)
    return None

def detect_broad_identity(text):
    return matches_any(BROAD_IDENTITY_KEYWORDS, text)

def context_is_overridden(text):
    return any(phrase_has_serving_context(p, text) for p in EXPANSION_PHRASES)

def context_is_historical(text):
    return any(phrase_has_serving_context(p, text) for p in HISTORICAL_PHRASES)

def context_is_aspirational(text):
    return matches_any(ASPIRATIONAL_PHRASES, text)

# =================================================
# TEXT EXTRACTION — concatenate across all 4 columns
# =================================================

def get_column_texts(row):
    texts = []
    for col in INPUT_COLS_PRIORITY:
        val = row.get(col, "")
        raw = normalize_text(val)
        # Commented out for alias check implementation.
        #raw = apply_aliases(raw)
        texts.append(raw)
    return texts

# ==========================
# MAIN CLASSIFICATION LOGIC
# =============================

def classify_row(row, taxonomy_entries):
    col_texts = get_column_texts(row)
    combined  = " ".join(t for t in col_texts if t.strip())

    # Extra notes (buzzword / cultural association mentions) get
    # appended to whatever flag this function ends up returning, no
    # matter which branch below resolves the classification.
    extra_notes = []

    def finalize(e1, e2, e3, flag):
        if extra_notes:
            note_text = "; ".join(extra_notes)
            flag = f"{flag}; {note_text}" if flag else note_text
        return (e1, e2, e3, flag)

    if not combined.strip():
        return finalize(GENERAL_POP, "", "", "Empty input")

    # Highest priority: context override / historical reference
    if context_is_overridden(combined):
        return finalize(GENERAL_POP, "", "", "Context override: expansion phrase detected")
    if context_is_historical(combined):
        return finalize(GENERAL_POP, "", "", "Context override: historical reference detected")

    aspirational = context_is_aspirational(combined)

    # "Cultural Association" often hides a specific named group (e.g.
    # "Kerala Cultural Association") -- always worth a second look,
    # regardless of how this row otherwise resolves.
    if re.search(r"\bcultural association\b", combined, re.IGNORECASE):
        extra_notes.append("'Cultural Association' detected - verify named group manually")

    # Case 11: grassroots / ambiguous equity words -- never suppress
    # silently now, always flagged whether or not a real signal is present.
    grassroots_state = check_grassroots_case(combined, taxonomy_entries) # Handle 'ethnocultural', 'marginalized' as well, since they have the same rule as 'grassroots'
    if grassroots_state == "no_signal":
        return finalize(GENERAL_POP, "", "", "Ambiguous equity term (e.g. grassroots, multicultural, refugee) with no paired ethnic signal")
    if grassroots_state == "has_signal":
        extra_notes.append("Equity/diversity buzzword present alongside a real signal - verify manually")

    # Cases 1-3: direct taxonomy matches, across FR columns
    all_matches = []
    for col_text in col_texts:
        if col_text.strip():
            all_matches.extend(find_taxonomy_matches(col_text, taxonomy_entries))

    seen = set()
    unique_matches = []
    for m in all_matches:
        if m["keyword"] not in seen:
            seen.add(m["keyword"])
            unique_matches.append(m)

    # Build a unified candidate pool: taxonomy + compound + pattern +
    # country + broad identity. Each candidate: (level1, level2, level3, depth, source)
    candidates = []
    for m in unique_matches:
        candidates.append((m["level1"], m["level2"] or "", m["level3"] or "",
                            m["depth"], "taxonomy"))

    # Afro-Caribbean / Afro-Latino -- add BOTH halves as candidates so
    # they combine into Multiple the same way two separate mentions would.
    for compound_pattern, group_tuples in ALWAYS_MULTIPLE_COMPOUNDS.items():
        if re.search(compound_pattern, combined, re.IGNORECASE):
            for (l1, l2, l3) in group_tuples:
                candidates.append((l1, l2, l3, 1, "compound"))

    pattern_result = find_pattern_match(combined)
    if pattern_result:
        l1, l2, l3, depth = pattern_result
        candidates.append((l1, l2, l3, depth, "pattern"))

    country_result = find_country_match(combined)
    if country_result:
        l1, l2, l3, depth = country_result
        candidates.append((l1, l2, l3, depth, "country"))

    # Broad identity labels with no specific taxonomy entry (Hispanic,
    # Latino, etc. -- NOT Black/Jewish/Arab, real taxonomy entries now,
    # already came through find_taxonomy_matches above). Folded into
    # the same pool so they can combine into Multiple alongside a
    # separately-named group too.
    if detect_broad_identity(combined):
        candidates.append((OTHER_ETHNIC, "", "", 1, "broad_identity"))

    # Dedupe by actual (level1, level2, level3) outcome; two different
    # detection methods (e.g. pattern rule + country map) landing on the
    # exact same conclusion is a single confirmed answer, not a multi-
    # group signal. Keeps the first-seen source label for the flag text.
    seen_outcomes = set()
    deduped_candidates = []
    for c in candidates:
        outcome_key = (c[0], c[1], c[2])
        if outcome_key not in seen_outcomes:
            seen_outcomes.add(outcome_key)
            deduped_candidates.append(c)
    candidates = deduped_candidates

    # Possible consulted-party mention (expert/advisor/biologist role)
    # rather than the population served -- flagged, never suppressed.
    if candidates and matches_any(EXPERT_ROLE_PHRASES, combined):
        extra_notes.append("Possible consulted-party mention (expert/advisor role) rather than served population - verify manually")

    # Case 9: BIPOC (context-aware) 
    # Checked here (after candidates are gathered) so we can tell whether
    # BIPOC is mentioned ALONGSIDE a real specific group (the ambiguous nuance from the README) vs. BIPOC being the only signal.
    bipoc_present = is_bipoc_real_target(combined)
    if bipoc_present:
        if candidates:
            # BIPOC + a specific named group both present. Exactly the
            # ambiguous case flagged in the README (e.g. "BIPOC visibility"
            # alongside "Asian Canadian artists"). Do not silently resolve;
            # flag for manual review.
            other_groups = sorted(set(c[0] for c in candidates))
            flag = ("Ambiguous: BIPOC mentioned alongside specific group(s) ("
                    + ", ".join(other_groups) + ") - verify manually")
        else:
            flag = "BIPOC signal detected"
        return finalize(MULTIPLE_ETHNIC, "", "", flag)

    # Resolve candidates: distinct Level 1 groups ANYWHERE in the pool
    # (not just tied at the deepest level) mean Multiple -- lets "African"
    # (depth 1) combine with "Black" (depth 2, a different L1 branch
    # under "Other Ethnic and Cultural Origins") instead of the deeper
    # one silently winning by depth alone.
    if candidates:
        distinct_l1 = set(c[0] for c in candidates)

        if len(distinct_l1) >= 2:
            flag = "Review: multiple distinct groups detected"
            if aspirational:
                flag += "; aspirational language present"
            return finalize(MULTIPLE_ETHNIC, "", "", flag)

        # All candidates share one Level 1 -- use depth to pick the most
        # specific. If several DIFFERENT branches still tie at the
        # deepest level within this one Level 1, that's still Multiple.
        max_depth = max(c[3] for c in candidates)
        deepest = [c for c in candidates if c[3] == max_depth]

        if len(deepest) >= 2:
            flag = "Review: multiple sub-groups within same origin"
            if aspirational:
                flag += "; aspirational language present"
            return finalize(MULTIPLE_ETHNIC, "", "", flag)

        l1, l2, l3, depth, source = deepest[0]
        flag = ""
        if source == "pattern":
            flag = "Pattern rule match (structured phrase)"
        elif source == "country":
            flag = "Country/nationality mapping match"
        elif source == "compound":
            flag = "Compound identity term match"
        elif source == "broad_identity":
            flag = "Broad identity term - review recommended"
        if aspirational:
            flag = (flag + "; " if flag else "") + "Review: aspirational language - future intent, not current population"
        return finalize(l1, l2, l3, flag)

    # Case 10: organization name lookup (LAST RESORT before General) NOTE: Need to highlight Black Canadian Women to flag
    org_result = check_org_name_lookup(combined)
    if org_result:
        l1, l2, l3 = org_result
        return finalize(l1, l2, l3, "Matched via known organization name lookup")

    # Case 12: fallback
    return finalize(GENERAL_POP, "", "", "")

# =====
# MAIN
# =====
def main():
    start_time = time.time()

    if len(sys.argv) < 3:
        print('Usage: python ethnic_taggerv3.py "C:\\Users\\oadode\\OneDrive - Edmonton Community Foundation\\Desktop\\Discretionary FR Scripting\\ECF-Discretionary-FR-Scripting\\Taxonomy - Definitions.xlsx" "C:\\Users\\oadode\\OneDrive - Edmonton Community Foundation\\Desktop\\Discretionary FR Scripting\\ECF-Discretionary-FR-Scripting\\FR testing.xlsx"')
        sys.exit(1)
 
    taxonomy_filepath = sys.argv[1]
    funding_filepath  = sys.argv[2]
 
    print(f"Loading taxonomy from: {taxonomy_filepath}")
    try:
        tax_df = pd.read_excel(taxonomy_filepath, sheet_name=TAXONOMY_SHEET, dtype=str)
    except Exception as e:
        print(f"Error loading taxonomy sheet '{TAXONOMY_SHEET}' from '{taxonomy_filepath}': {e}")
        sys.exit(1)
 
    print(f"Loading funding requests from: {funding_filepath}")
    try:
        data_df = pd.read_excel(funding_filepath, sheet_name=DATA_SHEET, dtype=str)
    except Exception as e:
        print(f"Error loading data sheet '{DATA_SHEET}' from '{funding_filepath}': {e}")
        sys.exit(1)
 
    print(f"Taxonomy rows: {len(tax_df)} | Data rows: {len(data_df)}")
 
    taxonomy_entries = build_taxonomy(tax_df)
    print(f"Taxonomy entries parsed: {len(taxonomy_entries)}")

    # Semantic Fallback - Level 2 Engine
    semantic_entries = None
    semantic_embeddings = None

    if SEMANTIC_AVAILABLE:
        scope_notes = semantic_fallback.build_scope_note_map(
            tax_df, TAXONOMY_ENTRY1, TAXONOMY_SCOPE_NOTES, safe_display)
        semantic_entries, semantic_texts = semantic_fallback.build_candidate_texts(
            taxonomy_entries, scope_notes)
        semantic_embeddings = semantic_fallback.get_taxonomy_embeddings(semantic_texts)
    else:
        print("semantic_fallback not available (sentence-transformers not installed) — skipping semantic suggestions.")
 
    for col in [OUTPUT_ETHNIC1, OUTPUT_ETHNIC2, OUTPUT_ETHNIC3, OUTPUT_FLAG, OUTPUT_SEMANTIC]:
        if col not in data_df.columns:
            data_df[col] = ""
    # Count how many rows fall into each outcome bucket, for summary stats at the end. Note that these are not mutually exclusive categories (e.g. a row with a pattern match that's also flagged for aspirational language would count in both "pattern" and "flagged"), but they give a general sense of how many hits came from each detection method and how many were flagged for review.
    stats = {"3-level": 0, "2-level": 0, "1-level": 0, "multiple": 0,
              "other": 0, "general": 0, "flagged": 0, "pattern": 0,
              "country": 0, "org_lookup": 0, "grassroots_filtered": 0, "semantic_suggested": 0}
 
    for idx, row in data_df.iterrows():
        # Initial deterministic engine (level 1) called from taxonomy definitions sheet
        e1, e2, e3, flag = classify_row(row, taxonomy_entries)
        data_df.at[idx, OUTPUT_ETHNIC1] = e1
        data_df.at[idx, OUTPUT_ETHNIC2] = e2
        data_df.at[idx, OUTPUT_ETHNIC3] = e3
        data_df.at[idx, OUTPUT_FLAG] = flag

        # Engine 2 - This is our Semantic fallback layer, which only triggers if Engine 1 returns General Population (i.e. no
        if SEMANTIC_AVAILABLE and e1 == GENERAL_POP:
            combined_text = " ".join(t for t in get_column_texts(row) if t.strip())
            suggestion = semantic_fallback.find_semantic_suggestion(
                combined_text, semantic_entries, semantic_embeddings)
            if suggestion:
                sl1, sl2, sl3, score, margin = suggestion
                parts = [p for p in [sl1, sl2, sl3] if p]
                data_df.at[idx, OUTPUT_SEMANTIC] = f"{' / '.join(parts)} (similarity: {score:.2f}, margin: {margin:.2f})"
                stats["semantic_suggested"] += 1
 
        if e1 == MULTIPLE_ETHNIC:
            stats["multiple"] += 1
        elif e1 == OTHER_ETHNIC:
            stats["other"] += 1
        elif e1 == GENERAL_POP:
            stats["general"] += 1
        elif e3:
            stats["3-level"] += 1
        elif e2:
            stats["2-level"] += 1
        else:
            stats["1-level"] += 1
 
        if "pattern rule" in flag.lower():
            stats["pattern"] += 1
        if "country" in flag.lower():
            stats["country"] += 1
        if "organization name" in flag.lower():
            stats["org_lookup"] += 1
        if "ambiguous equity" in flag.lower():
            stats["grassroots_filtered"] += 1
        if flag:
            stats["flagged"] += 1
 
    wb = load_workbook(funding_filepath)
    ws = wb[DATA_SHEET]
    headers = {cell.value: cell.column for cell in ws[1]}
 
    for col_name in [OUTPUT_ETHNIC1, OUTPUT_ETHNIC2, OUTPUT_ETHNIC3, OUTPUT_FLAG, OUTPUT_SEMANTIC]:
        if col_name not in headers:
            new_col = ws.max_column + 1
            ws.cell(row=1, column=new_col, value=col_name)
            headers[col_name] = new_col
 
    for i, (idx, row) in enumerate(data_df.iterrows(), start=2):
        ws.cell(row=i, column=headers[OUTPUT_ETHNIC1], value=data_df.at[idx, OUTPUT_ETHNIC1])
        ws.cell(row=i, column=headers[OUTPUT_ETHNIC2], value=data_df.at[idx, OUTPUT_ETHNIC2])
        ws.cell(row=i, column=headers[OUTPUT_ETHNIC3], value=data_df.at[idx, OUTPUT_ETHNIC3])
        ws.cell(row=i, column=headers[OUTPUT_FLAG],    value=data_df.at[idx, OUTPUT_FLAG])
        ws.cell(row=i, column=headers[OUTPUT_SEMANTIC], value=data_df.at[idx, OUTPUT_SEMANTIC]) # Writes "" for every row that didn't receive a suggestion (General pop.)
    wb.save(funding_filepath)
    
    print("\nResults:")
    for k, v in stats.items():
        print(f"  {k}: {v}")
        # print(f"  {k}: {v}; Column #: {get_column_letter(headers[OUTPUT_ETHNIC1])} - {get_column_letter(headers[OUTPUT_ETHNIC3])}, Flag: {get_column_letter(headers[OUTPUT_FLAG])}")

    print(f"\nOutput written to: {funding_filepath}")
    elapsed = time.time() - start_time
    print(f"Ethnic classification completed in {elapsed:.1f} seconds.")
 
if __name__ == "__main__":
    main()